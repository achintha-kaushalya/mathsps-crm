#!/usr/bin/env python3
"""
Fixes dates in Master_Leads_GoogleSheets.xlsx to be actual datetime objects
instead of strings, so that Google Sheets can perform numeric date filtering.
"""

import openpyxl
from datetime import datetime

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Master_Leads_GoogleSheets.xlsx'

# Load workbook
wb = openpyxl.load_workbook(DST)
ws = wb['Master Leads']

print("Converting date strings to real Date objects...")

# Process rows in Master Leads
for r in range(2, ws.max_row + 1):
    cell = ws.cell(r, 5) # Column E: Date Added
    val = cell.value
    if val:
        # If it's a string, convert to datetime.date
        if isinstance(val, str):
            val_str = val.strip()
            # Try parsing YYYY-MM-DD
            try:
                dt = datetime.strptime(val_str, "%Y-%m-%d").date()
                cell.value = dt
                cell.number_format = "YYYY-MM-DD"
                print(f"Row {r}: Converted string '{val_str}' to Date object")
            except ValueError:
                # Try parsing with time if present
                try:
                    dt = datetime.strptime(val_str, "%Y-%m-%d %H:%M:%S").date()
                    cell.value = dt
                    cell.number_format = "YYYY-MM-DD"
                    print(f"Row {r}: Converted string '{val_str}' to Date object")
                except ValueError:
                    print(f"Row {r}: Could not parse '{val_str}' as date")
        elif isinstance(val, (int, float)):
            # If it's a number, keep it as a number but ensure date format
            cell.number_format = "YYYY-MM-DD"
            print(f"Row {r}: Numeric value {val} kept, format set to YYYY-MM-DD")
        else:
            # It's already a datetime/date object
            cell.number_format = "YYYY-MM-DD"
            print(f"Row {r}: Already a {type(val)}")

# Save workbook
wb.save(DST)
print("Saved fixed workbook successfully!")
