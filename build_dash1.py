#!/usr/bin/env python3
"""
Lead CRM - Dashboard 1: Executive Dashboard Builder
Builds a professional Excel Executive Dashboard referencing Master Leads.
"""

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
SRC = r'c:\Users\achir\Desktop\Lead tracking PS\Lead_Management_System_v8_2 (1).xlsx'
DST = r'c:\Users\achir\Desktop\Lead tracking PS\Lead_CRM_Dashboards.xlsx'
ML  = "'Master Leads'"   # Cross-sheet formula prefix

# ── Color Palette ─────────────────────────────────────────────────────────────
NAVY     = "0D2744"
BLUE     = "1565C0"
ROYAL    = "1976D2"
SKY      = "2196F3"
ICE      = "DBEAFE"
WHITE    = "FFFFFF"
OFF_W    = "F0F4F8"
LGRAY    = "E2E8F0"
GRAY     = "64748B"
DGRAY    = "1E293B"
GREEN    = "166534"
LGREEN   = "DCFCE7"
AMBER    = "C2410C"
LAMBER   = "FFF7ED"
RED      = "991B1B"
LRED     = "FEE2E2"
PURPLE   = "581C87"
LPURPLE  = "F3E8FF"
TEAL     = "134E4A"
LTEAL    = "CCFBF1"
INDIGO   = "312E81"
LINDIGO  = "E0E7FF"
CARD_BG  = "FFFFFF"
CHART_ACCENT1 = "3B82F6"
CHART_ACCENT2 = "10B981"
CHART_ACCENT3 = "F59E0B"
CHART_ACCENT4 = "EF4444"
CHART_ACCENT5 = "8B5CF6"
CHART_ACCENT6 = "06B6D4"

# ── Style Helpers ─────────────────────────────────────────────────────────────
def pf(color):
    return PatternFill("solid", fgColor=color)

def ft(size=10, bold=False, color="000000", italic=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def side(style=None, color="CCCCCC"):
    return Side(border_style=style, color=color) if style else Side(border_style=None)

def bd(left=None, right=None, top=None, bottom=None, lc="CCCCCC", rc="CCCCCC", tc="CCCCCC", bc="CCCCCC"):
    return Border(
        left=side(left, lc), right=side(right, rc),
        top=side(top, tc),   bottom=side(bottom, bc)
    )

def merge_set(ws, r1, c1, r2, c2, val=None, fill_color=None,
              font=None, align=None, numfmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if val is not None:   cell.value = val
    if font:              cell.font  = font
    if align:             cell.alignment = align
    if numfmt:            cell.number_format = numfmt
    if fill_color:
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).fill = pf(fill_color)
    return cell

# ── KPI Card Drawer ───────────────────────────────────────────────────────────
def kpi_card(ws, row, col, title, formula, sublabel, accent, numfmt=None):
    """
    Draws a 3-row × 4-col KPI card.
      Row 0 : colored header strip with title
      Row 1 : big number (formula)
      Row 2 : small sublabel text
    """
    W = 4   # card width in cols
    bg = CARD_BG

    # ── Fill cells ───────────────────────────────────────────────────────────
    for r in range(row, row + 3):
        for c in range(col, col + W):
            ws.cell(r, c).fill = pf(accent if r == row else bg)

    # ── Outer card border (thin blue-gray) ────────────────────────────────────
    def apply_border(r, c):
        l = side("medium", accent)  if c == col      else side(None)
        ri= side("thin",   "D1D5DB") if c == col+W-1 else side(None)
        t = side("thin",   "D1D5DB") if r == row      else side(None)
        b = side("thin",   "D1D5DB") if r == row+2    else side(None)
        ws.cell(r, c).border = Border(left=l, right=ri, top=t, bottom=b)

    for r in range(row, row + 3):
        for c in range(col, col + W):
            apply_border(r, c)

    # ── Header strip: title ───────────────────────────────────────────────────
    merge_set(ws, row, col, row, col+W-1,
              val=title.upper(),
              fill_color=accent,
              font=ft(8, True, WHITE),
              align=al("left", "center"))

    # ── Value row ─────────────────────────────────────────────────────────────
    vc = merge_set(ws, row+1, col, row+1, col+W-1,
                   val=formula,
                   fill_color=bg,
                   font=ft(22, True, accent),
                   align=al("left", "center"))
    if numfmt:
        vc.number_format = numfmt

    # ── Sublabel ─────────────────────────────────────────────────────────────
    merge_set(ws, row+2, col, row+2, col+W-1,
              val=sublabel,
              fill_color=bg,
              font=ft(8, False, GRAY),
              align=al("left", "top"))

# ── Section Header ────────────────────────────────────────────────────────────
def section_hdr(ws, row, c1, c2, label):
    for c in range(1, c2+2):
        ws.cell(row, c).fill = pf(LGRAY)
    merge_set(ws, row, c1, row, c2,
              val=label,
              fill_color=LGRAY,
              font=ft(10, True, DGRAY),
              align=al("left", "center"))

# ═════════════════════════════════════════════════════════════════════════════
# SETUP
# ═════════════════════════════════════════════════════════════════════════════
shutil.copy2(SRC, DST)
wb = load_workbook(DST)

# ── Hidden helper data sheet ──────────────────────────────────────────────────
HELPER = "_DB1_Data"
if HELPER in wb.sheetnames:
    del wb[HELPER]
hws = wb.create_sheet(HELPER)
hws.sheet_state = "hidden"

# ── Dashboard sheet ───────────────────────────────────────────────────────────
DNAME = "Executive Dashboard"
if DNAME in wb.sheetnames:
    del wb[DNAME]
ws = wb.create_sheet(DNAME, 0)
ws.sheet_properties.tabColor = BLUE
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 85

# ═════════════════════════════════════════════════════════════════════════════
# COLUMN & ROW SIZING
# Cols: A(margin)|B-E(card1)|F(gap)|G-J(card2)|K(gap)|L-O(card3)|
#       P(gap)|Q-T(card4)|U(gap)|V-Y(card5)|Z(margin)
# ═════════════════════════════════════════════════════════════════════════════
COL_CARD = 3.8   # card column width
COL_GAP  = 1.2   # gap between cards
COL_MARG = 1.5   # side margins

ws.column_dimensions["A"].width = COL_MARG
ws.column_dimensions["Z"].width = COL_MARG

CARD_COLS = [2,3,4,5,  7,8,9,10,  12,13,14,15,  17,18,19,20,  22,23,24,25]
GAP_COLS  = [6, 11, 16, 21]

for c in CARD_COLS:
    ws.column_dimensions[get_column_letter(c)].width = COL_CARD
for c in GAP_COLS:
    ws.column_dimensions[get_column_letter(c)].width = COL_GAP

for c in range(28, 70):
    ws.column_dimensions[get_column_letter(c)].width = 13

# Row heights
ROW_H = {
    1:  7,    # top margin
    2:  52,   # Title
    3:  20,   # subtitle bar
    4:  12,   # spacer
    5:  14,   # KPI row1 header strip
    6:  38,   # KPI row1 value
    7:  15,   # KPI row1 sublabel
    8:  10,   # spacer
    9:  14,   # KPI row2 header strip
    10: 38,   # KPI row2 value
    11: 15,   # KPI row2 sublabel
    12: 14,   # spacer
    13: 26,   # section header
}
for r, h in ROW_H.items():
    ws.row_dimensions[r].height = h
for r in range(14, 90):
    ws.row_dimensions[r].height = 14

# ═════════════════════════════════════════════════════════════════════════════
# BACKGROUND
# ═════════════════════════════════════════════════════════════════════════════
for r in range(1, 85):
    for c in range(1, 27):
        ws.cell(r, c).fill = pf(OFF_W)

# ═════════════════════════════════════════════════════════════════════════════
# TITLE BANNER
# ═════════════════════════════════════════════════════════════════════════════
for c in range(1, 27):
    ws.cell(2, c).fill = pf(NAVY)
merge_set(ws, 2, 2, 2, 25,
    val="  LEAD MANAGEMENT SYSTEM   |   EXECUTIVE DASHBOARD",
    fill_color=NAVY,
    font=ft(20, True, WHITE, name="Calibri"),
    align=al("left", "center"))

for c in range(1, 27):
    ws.cell(3, c).fill = pf(ROYAL)
merge_set(ws, 3, 2, 3, 25,
    val="  Live data from Master Leads   \u00b7   Auto-refreshes with new entries   \u00b7   Professional CRM Analytics",
    fill_color=ROYAL,
    font=ft(9, False, "BFDBFE"),
    align=al("left", "center"))

# ═════════════════════════════════════════════════════════════════════════════
# KPI CARDS - ROW 1 (rows 5-7)
# ═════════════════════════════════════════════════════════════════════════════
kpis_r1 = [
    # (start_col, title, formula, sublabel, accent, numfmt)
    (2,  "TOTAL LEADS",
     f"=COUNTA({ML}!B:B)-1",
     "All entries in Master Leads",
     NAVY, None),

    (7,  "PAID STUDENTS",
     f'=COUNTIF({ML}!O:O,"Yes")',
     "Confirmed fee paid (col O = Yes)",
     GREEN, None),

    (12, "CONVERSION RATE",
     f'=IFERROR(COUNTIF({ML}!F:F,"Converted")/(COUNTA({ML}!B:B)-1),0)',
     "Converted leads \u00f7 Total leads",
     TEAL, "0.0%"),

    (17, "REPEAT STUDENTS",
     f'=COUNTIF({ML}!M:M,"Yes*")',
     "Same student, multiple campaigns",
     PURPLE, None),

    (22, "ACTIVE LEADS",
     f'=COUNTIF({ML}!F:F,"New")+COUNTIF({ML}!F:F,"Contacted")',
     "Status = New or Contacted",
     BLUE, None),
]

for col, title, formula, sublabel, accent, numfmt in kpis_r1:
    kpi_card(ws, 5, col, title, formula, sublabel, accent, numfmt)

# ═════════════════════════════════════════════════════════════════════════════
# KPI CARDS - ROW 2 (rows 9-11)
# ═════════════════════════════════════════════════════════════════════════════
kpis_r2 = [
    (2,  "NO ANSWER",
     f'=COUNTIF({ML}!F:F,"No Answer")',
     "Calls not picked up",
     AMBER, None),

    (7,  "CONVERTED",
     f'=COUNTIF({ML}!F:F,"Converted")',
     "Successfully converted",
     GREEN, None),

    (12, "INTERESTED",
     f'=COUNTIF({ML}!F:F,"Interested")',
     "Interested but not yet paid",
     ROYAL, None),

    (17, "TODAY'S LEADS",
     f"=COUNTIF({ML}!E:E,TODAY())",
     "New leads added today",
     INDIGO, None),

    (22, "THIS MONTH",
     f'=COUNTIFS({ML}!E:E,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
     f'{ML}!E:E,"<="&EOMONTH(TODAY(),0))',
     "Leads added this month",
     TEAL, None),
]

for col, title, formula, sublabel, accent, numfmt in kpis_r2:
    kpi_card(ws, 9, col, title, formula, sublabel, accent, numfmt)

# ═════════════════════════════════════════════════════════════════════════════
# SECTION HEADER - Charts
# ═════════════════════════════════════════════════════════════════════════════
section_hdr(ws, 13, 2, 25, "  PERFORMANCE ANALYTICS   \u2014   Charts auto-update from Master Leads")

# ═════════════════════════════════════════════════════════════════════════════
# HELPER DATA SHEET - Chart Data Sources
# ═════════════════════════════════════════════════════════════════════════════

def hdr(hws, r, c, val):
    hws.cell(r, c).value = val
    hws.cell(r, c).font  = Font(name="Calibri", size=9, bold=True, color=NAVY)

def dat(hws, r, c, val, numfmt=None):
    hws.cell(r, c).value = val
    hws.cell(r, c).font  = Font(name="Calibri", size=9)
    if numfmt:
        hws.cell(r, c).number_format = numfmt

# ── Block A: Monthly Trend (rows 1-13) ────────────────────────────────────────
hdr(hws, 1, 1, "Month");  hdr(hws, 1, 2, "Leads")
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for i, m in enumerate(months):
    mn = i + 1
    dat(hws, 2+i, 1, m)
    dat(hws, 2+i, 2,
        f"=COUNTIFS('{ML.strip(chr(39))}'" + f"!E:E,"
        f'">="&DATE(YEAR(TODAY()),{mn},1),'
        f"'{ML.strip(chr(39))}'"
        f"!E:E,\"<=\"&EOMONTH(DATE(YEAR(TODAY()),{mn},1),0))")

# Fix the formula (ML already has quotes)
for i, m in enumerate(months):
    mn = i + 1
    hws.cell(2+i, 2).value = (
        f'=COUNTIFS({ML}!E:E,">="&DATE(YEAR(TODAY()),{mn},1),'
        f'{ML}!E:E,"<="&EOMONTH(DATE(YEAR(TODAY()),{mn},1),0))'
    )

# ── Block B: Status Distribution (rows 15-23) ─────────────────────────────────
hdr(hws, 15, 1, "Status"); hdr(hws, 15, 2, "Count")
statuses = ["New","Contacted","Converted","Not Interested",
            "No Answer","Interested","Follow-up","Second Call Done","Not Picked Up"]
for i, s in enumerate(statuses):
    dat(hws, 16+i, 1, s)
    hws.cell(16+i, 2).value = f'=COUNTIF({ML}!F:F,"{s}")'

# ── Block C: Paid vs Unpaid (rows 26-28) ─────────────────────────────────────
hdr(hws, 26, 1, "Payment"); hdr(hws, 26, 2, "Count")
dat(hws, 27, 1, "Paid")
hws.cell(27, 2).value = f'=COUNTIF({ML}!O:O,"Yes")'
dat(hws, 28, 1, "Unpaid")
hws.cell(28, 2).value = f'=MAX(0,COUNTA({ML}!B:B)-1-COUNTIF({ML}!O:O,"Yes"))'

# ── Block D: Campaign Performance (rows 30-40) ────────────────────────────────
hdr(hws, 30, 1, "Campaign"); hdr(hws, 30, 2, "Leads"); hdr(hws, 30, 3, "Paid")
campaigns = ["B1","B2","B3","B4","B5","B6","B7","B8","B9","B10"]
for i, camp in enumerate(campaigns):
    dat(hws, 31+i, 1, camp)
    hws.cell(31+i, 2).value = f'=COUNTIF({ML}!L:L,"{camp}")'
    hws.cell(31+i, 3).value = f'=COUNTIFS({ML}!L:L,"{camp}",{ML}!O:O,"Yes")'

# ── Block E: Daily Trend - Last 7 Days (rows 42-49) ──────────────────────────
hdr(hws, 42, 1, "Date"); hdr(hws, 42, 2, "Leads")
for i in range(7):
    offset = 6 - i   # 6,5,4,3,2,1,0 days ago
    hws.cell(43+i, 1).value = f"=TODAY()-{offset}"
    hws.cell(43+i, 1).number_format = "DD-MMM"
    hws.cell(43+i, 2).value = f'=COUNTIF({ML}!E:E,TODAY()-{offset})'

# ═════════════════════════════════════════════════════════════════════════════
# CHARTS
# ═════════════════════════════════════════════════════════════════════════════

# ─── Chart 1: Monthly Leads Trend (Line) ─────────────────────────────────────
line1 = LineChart()
line1.title  = "Monthly Leads Trend (2026)"
line1.style  = 10
line1.y_axis.title = "Number of Leads"
line1.x_axis.title = "Month"
line1.legend = None

cats1 = Reference(hws, min_col=1, max_col=1, min_row=2, max_row=13)
vals1 = Reference(hws, min_col=2, max_col=2, min_row=1, max_row=13)
line1.add_data(vals1, titles_from_data=True)
line1.set_categories(cats1)

s = line1.series[0]
s.graphicalProperties.line.solidFill = CHART_ACCENT1
s.graphicalProperties.line.width = 28000
s.marker.symbol = "circle"
s.marker.size   = 7
s.marker.graphicalProperties.solidFill   = CHART_ACCENT1
s.marker.graphicalProperties.line.solidFill = WHITE
s.smooth = True

line1.width  = 21
line1.height = 10
ws.add_chart(line1, "B15")

# ─── Chart 2: Lead Status Pie ─────────────────────────────────────────────────
pie2 = PieChart()
pie2.title  = "Lead Status Distribution"
pie2.style  = 10

cats2 = Reference(hws, min_col=1, max_col=1, min_row=16, max_row=24)
vals2 = Reference(hws, min_col=2, max_col=2, min_row=15, max_row=24)
pie2.add_data(vals2, titles_from_data=True)
pie2.set_categories(cats2)

pie_colors = [BLUE, ROYAL, GREEN, RED, AMBER, CHART_ACCENT1, PURPLE, TEAL, CHART_ACCENT3]
for idx, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    pie2.series[0].dPt.append(pt)

pie2.width  = 12
pie2.height = 10
ws.add_chart(pie2, "L15")

# ─── Chart 3: Paid vs Unpaid Doughnut ─────────────────────────────────────────
donut3 = DoughnutChart()
donut3.title  = "Paid vs Unpaid"
donut3.style  = 10
donut3.holeSize = 55

cats3 = Reference(hws, min_col=1, max_col=1, min_row=27, max_row=28)
vals3 = Reference(hws, min_col=2, max_col=2, min_row=26, max_row=28)
donut3.add_data(vals3, titles_from_data=True)
donut3.set_categories(cats3)

pt_paid   = DataPoint(idx=0); pt_paid.graphicalProperties.solidFill   = GREEN
pt_unpaid = DataPoint(idx=1); pt_unpaid.graphicalProperties.solidFill = LGRAY
donut3.series[0].dPt.extend([pt_paid, pt_unpaid])

donut3.width  = 10
donut3.height = 10
ws.add_chart(donut3, "S15")

# ─── Chart 4: Campaign Performance Bar ────────────────────────────────────────
bar4 = BarChart()
bar4.type   = "bar"
bar4.title  = "Campaign Performance"
bar4.style  = 10
bar4.y_axis.title = "Campaign"
bar4.x_axis.title = "Count"
bar4.grouping = "clustered"

cats4       = Reference(hws, min_col=1, max_col=1, min_row=31, max_row=40)
vals4_leads = Reference(hws, min_col=2, max_col=2, min_row=30, max_row=40)
vals4_paid  = Reference(hws, min_col=3, max_col=3, min_row=30, max_row=40)
bar4.add_data(vals4_leads, titles_from_data=True)
bar4.add_data(vals4_paid,  titles_from_data=True)
bar4.set_categories(cats4)

bar4.series[0].graphicalProperties.solidFill = CHART_ACCENT1
bar4.series[1].graphicalProperties.solidFill = CHART_ACCENT2

bar4.width  = 16
bar4.height = 10
ws.add_chart(bar4, "B35")

# ─── Chart 5: Daily Leads Trend (Last 7 Days) ────────────────────────────────
line5 = LineChart()
line5.title  = "Daily Leads \u2014 Last 7 Days"
line5.style  = 10
line5.y_axis.title = "Leads"
line5.x_axis.title = "Date"
line5.legend = None

cats5 = Reference(hws, min_col=1, max_col=1, min_row=43, max_row=49)
vals5 = Reference(hws, min_col=2, max_col=2, min_row=42, max_row=49)
line5.add_data(vals5, titles_from_data=True)
line5.set_categories(cats5)

s5 = line5.series[0]
s5.graphicalProperties.line.solidFill = CHART_ACCENT3
s5.graphicalProperties.line.width = 28000
s5.marker.symbol = "diamond"
s5.marker.size   = 8
s5.marker.graphicalProperties.solidFill   = CHART_ACCENT3
s5.marker.graphicalProperties.line.solidFill = WHITE
s5.smooth = True

line5.width  = 16
line5.height = 10
ws.add_chart(line5, "P35")

# ═════════════════════════════════════════════════════════════════════════════
# FOOTER / NOTES
# ═════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[57].height = 18
for c in range(1, 27):
    ws.cell(57, c).fill = pf(LGRAY)
merge_set(ws, 57, 2, 57, 25,
    val="  NOTE: Campaign chart uses B1-B10 by default. Update campaign names in _DB1_Data sheet (col A, rows 31-40) to match your actual campaigns.",
    fill_color=LGRAY,
    font=ft(8, False, GRAY),
    align=al("left", "center"))

# ═════════════════════════════════════════════════════════════════════════════
# FREEZE PANES  &  SHEET ORDER
# ═════════════════════════════════════════════════════════════════════════════
ws.freeze_panes = "B4"

# Move Executive Dashboard first, _DB1_Data last
wb.move_sheet(DNAME, offset=-wb.index(wb[DNAME]))

# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
wb.save(DST)
print(f"Dashboard 1 saved  ->  {DST}")
print("Sheets:", wb.sheetnames)
