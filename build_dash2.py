#!/usr/bin/env python3
"""
Lead CRM - Dashboard 2: Grade Interest Dashboard
Shows lead breakdown by grade with charts and percentages.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Lead_CRM_Dashboards.xlsx'
ML  = "'Master Leads'"

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY      = "0D2744"; BLUE     = "1565C0"; ROYAL    = "1976D2"
WHITE     = "FFFFFF"; OFF_W    = "F0F4F8"; LGRAY    = "E2E8F0"
GRAY      = "64748B"; DGRAY    = "1E293B"; GREEN    = "166534"
LGREEN    = "DCFCE7"; TEAL     = "134E4A"; LTEAL    = "CCFBF1"
PURPLE    = "581C87"; LPURPLE  = "F3E8FF"; AMBER    = "C2410C"
LAMBER    = "FFF7ED"; RED      = "991B1B"; LRED     = "FEE2E2"
INDIGO    = "312E81"; LINDIGO  = "E0E7FF"; CARD_BG  = "FFFFFF"

GRADE_COLORS = {
    "6":  "1D4ED8",   # blue
    "7":  "0369A1",   # sky
    "8":  "047857",   # green
    "9":  "7C3AED",   # purple
    "10": "B45309",   # amber
    "11": "B91C1C",   # red
}

def pf(color):   return PatternFill("solid", fgColor=color)
def ft(size=10, bold=False, color="000000", italic=False, name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def side(style=None, color="CCCCCC"):
    return Side(border_style=style, color=color) if style else Side(border_style=None)
def bd_all(color="D1D5DB", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def merge_set(ws, r1, c1, r2, c2, val=None, fill_color=None,
              font=None, align=None, numfmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if val is not None:   cell.value = val
    if font:              cell.font  = font
    if align:             cell.alignment = align
    if numfmt:            cell.number_format = numfmt
    if fill_color:
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).fill = pf(fill_color)
    return cell

def section_hdr(ws, row, c1, c2, label, bg=LGRAY):
    for c in range(1, c2+2):
        ws.cell(row, c).fill = pf(bg)
    merge_set(ws, row, c1, row, c2, val=label,
              fill_color=bg, font=ft(10, True, DGRAY), align=al("left", "center"))

# ═════════════════════════════════════════════════════════════════════════════
# LOAD & ADD SHEET
# ═════════════════════════════════════════════════════════════════════════════
wb = load_workbook(DST)

HELPER = "_DB2_Data"
if HELPER in wb.sheetnames: del wb[HELPER]
hws = wb.create_sheet(HELPER)
hws.sheet_state = "hidden"

DNAME = "Grade Interest"
if DNAME in wb.sheetnames: del wb[DNAME]
ws = wb.create_sheet(DNAME, 1)   # insert after Executive Dashboard
ws.sheet_properties.tabColor = BLUE
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 85

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 1.5
for c in range(2, 32):
    ws.column_dimensions[get_column_letter(c)].width = 3.8
for c in range(32, 70):
    ws.column_dimensions[get_column_letter(c)].width = 13

# ── Row heights ───────────────────────────────────────────────────────────────
for r, h in {1:7, 2:52, 3:20, 4:14, 5:26, 6:14}.items():
    ws.row_dimensions[r].height = h
for r in range(7, 90):
    ws.row_dimensions[r].height = 14

# ── Background ────────────────────────────────────────────────────────────────
for r in range(1, 85):
    for c in range(1, 30):
        ws.cell(r, c).fill = pf(OFF_W)

# ═════════════════════════════════════════════════════════════════════════════
# TITLE
# ═════════════════════════════════════════════════════════════════════════════
for c in range(1, 30): ws.cell(2, c).fill = pf(NAVY)
merge_set(ws, 2, 2, 2, 28,
    val="  GRADE INTEREST DASHBOARD   |   Leads by Grade (6 - 11)",
    fill_color=NAVY, font=ft(20, True, WHITE), align=al("left", "center"))

for c in range(1, 30): ws.cell(3, c).fill = pf(ROYAL)
merge_set(ws, 3, 2, 3, 28,
    val="  Shows how many leads enquired for each grade  \u00b7  Auto-updates with new data",
    fill_color=ROYAL, font=ft(9, False, "BFDBFE"), align=al("left", "center"))

# ═════════════════════════════════════════════════════════════════════════════
# HELPER DATA (hws)
# ═════════════════════════════════════════════════════════════════════════════
# Block A: Grade breakdown (rows 1-8)
hws.cell(1, 1).value = "Grade"; hws.cell(1, 1).font = Font(bold=True)
hws.cell(1, 2).value = "Leads"; hws.cell(1, 2).font = Font(bold=True)
hws.cell(1, 3).value = "Paid";  hws.cell(1, 3).font = Font(bold=True)

GRADES = ["6","7","8","9","10","11"]
TOTAL_FORMULA = f"=COUNTA({ML}!B:B)-1"

for i, g in enumerate(GRADES):
    row = 2 + i
    hws.cell(row, 1).value = f"Grade {g}"
    hws.cell(row, 2).value = f'=COUNTIF({ML}!G:G,{g})'
    hws.cell(row, 3).value = f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'

# Block B: Monthly by Grade (rows 11-24 = 12 months per grade)
# Store month-grade matrix for trend: row=month, col offset per grade
hws.cell(10, 1).value = "Month"
for gi, g in enumerate(GRADES):
    hws.cell(10, 2+gi).value = f"Grade {g}"
    hws.cell(10, 2+gi).font = Font(bold=True)

months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for mi, m in enumerate(months):
    mn = mi + 1
    hws.cell(11+mi, 1).value = m
    for gi, g in enumerate(GRADES):
        hws.cell(11+mi, 2+gi).value = (
            f'=COUNTIFS({ML}!G:G,{g},'
            f'{ML}!E:E,">="&DATE(YEAR(TODAY()),{mn},1),'
            f'{ML}!E:E,"<="&EOMONTH(DATE(YEAR(TODAY()),{mn},1),0))'
        )

# ═════════════════════════════════════════════════════════════════════════════
# GRADE SUMMARY TABLE on Dashboard
# ═════════════════════════════════════════════════════════════════════════════
section_hdr(ws, 5, 2, 28, "  GRADE BREAKDOWN   —   Leads, Paid Students & Percentage per Grade")

# Table starts row 7
# Headers
header_cols = ["Grade", "Total Leads", "Paid Students",
               "Lead %", "Paid %", "Interested \u2192 Paid Conv."]
h_widths    = [3, 4, 4, 3, 3, 5]   # col span per header
h_colors    = [NAVY, BLUE, GREEN, TEAL, PURPLE, AMBER]

col = 2
for i, (hdr_text, span, hcol) in enumerate(zip(header_cols, h_widths, h_colors)):
    merge_set(ws, 7, col, 7, col+span-1,
              val=hdr_text, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

ws.row_dimensions[7].height = 22

# Grade rows 8-13
grade_acc = [GRADE_COLORS[g] for g in GRADES]
total_leads_cell = f"COUNTA({ML}!B:B)-1"

col_positions = [2, 5, 9, 13, 16, 19]  # start col for each column

for gi, g in enumerate(GRADES):
    r = 8 + gi
    ws.row_dimensions[r].height = 20

    bg = LGREEN if gi % 2 == 0 else WHITE
    acc = grade_acc[gi]

    # Fill entire row
    for c in range(2, 25):
        ws.cell(r, c).fill = pf(bg)
        ws.cell(r, c).border = bd_all()

    # Grade label
    merge_set(ws, r, 2, r, 4,
              val=f"Grade {g}", fill_color=acc,
              font=ft(10, True, WHITE), align=al("center", "center"))

    # Total leads
    leads_formula = f"=COUNTIF({ML}!G:G,{g})"
    merge_set(ws, r, 5, r, 8,
              val=leads_formula, fill_color=bg,
              font=ft(11, True, DGRAY), align=al("center", "center"))

    # Paid
    paid_formula = f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'
    merge_set(ws, r, 9, r, 12,
              val=paid_formula, fill_color=bg,
              font=ft(11, True, GREEN), align=al("center", "center"))

    # Lead %
    lead_pct = f'=IFERROR(COUNTIF({ML}!G:G,{g})/({total_leads_cell}),0)'
    c1 = merge_set(ws, r, 13, r, 15,
              val=lead_pct, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c1.number_format = "0.0%"

    # Paid %
    paid_pct = f'=IFERROR(COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")/COUNTIF({ML}!G:G,{g}),0)'
    c2 = merge_set(ws, r, 16, r, 18,
              val=paid_pct, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c2.number_format = "0.0%"

    # Interested -> Paid conversion
    conv_formula = (
        f'=IFERROR(COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")/'
        f'(COUNTIFS({ML}!G:G,{g},{ML}!F:F,"Interested")+COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")),0)'
    )
    c3 = merge_set(ws, r, 19, r, 23,
              val=conv_formula, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c3.number_format = "0.0%"

# TOTAL row
r_total = 14
ws.row_dimensions[r_total].height = 22
for c in range(2, 24):
    ws.cell(r_total, c).fill = pf(NAVY)
    ws.cell(r_total, c).border = bd_all(NAVY)

merge_set(ws, r_total, 2, r_total, 4,
    val="TOTAL", fill_color=NAVY, font=ft(10, True, WHITE), align=al("center", "center"))
merge_set(ws, r_total, 5, r_total, 8,
    val=f"={total_leads_cell}", fill_color=NAVY,
    font=ft(11, True, WHITE), align=al("center", "center"))
merge_set(ws, r_total, 9, r_total, 12,
    val=f'=COUNTIF({ML}!O:O,"Yes")', fill_color=NAVY,
    font=ft(11, True, WHITE), align=al("center", "center"))
merge_set(ws, r_total, 13, r_total, 23,
    val="100%", fill_color=NAVY, font=ft(10, True, WHITE), align=al("center", "center"))

# ═════════════════════════════════════════════════════════════════════════════
# CHARTS
# ═════════════════════════════════════════════════════════════════════════════

# Chart header
ws.row_dimensions[16].height = 26
section_hdr(ws, 16, 2, 28, "  GRADE CHARTS   —   Visual breakdown of leads by grade")

# ─── Chart A: Pie – Grade Distribution ────────────────────────────────────────
pie_a = PieChart()
pie_a.title = "Grade Distribution (All Leads)"
pie_a.style = 10

cats_a = Reference(hws, min_col=1, max_col=1, min_row=2, max_row=7)
vals_a = Reference(hws, min_col=2, max_col=2, min_row=1, max_row=7)
pie_a.add_data(vals_a, titles_from_data=True)
pie_a.set_categories(cats_a)

for idx, g in enumerate(GRADES):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = GRADE_COLORS[g]
    pie_a.series[0].dPt.append(pt)

pie_a.width  = 12
pie_a.height = 11
ws.add_chart(pie_a, "B17")

# ─── Chart B: Horizontal Bar – Leads by Grade ─────────────────────────────────
bar_b = BarChart()
bar_b.type     = "bar"
bar_b.title    = "Leads by Grade"
bar_b.style    = 10
bar_b.y_axis.title = "Grade"
bar_b.x_axis.title = "Number of Leads"
bar_b.grouping = "clustered"
bar_b.legend   = None

cats_b = Reference(hws, min_col=1, max_col=1, min_row=2, max_row=7)
vals_b = Reference(hws, min_col=2, max_col=2, min_row=1, max_row=7)
bar_b.add_data(vals_b, titles_from_data=True)
bar_b.set_categories(cats_b)
bar_b.series[0].graphicalProperties.solidFill = "1565C0"

bar_b.width  = 13
bar_b.height = 11
ws.add_chart(bar_b, "L17")

# ─── Chart C: Column – Leads vs Paid by Grade ─────────────────────────────────
col_c = BarChart()
col_c.type     = "col"
col_c.title    = "Leads vs Paid by Grade"
col_c.style    = 10
col_c.x_axis.title = "Grade"
col_c.y_axis.title = "Count"
col_c.grouping = "clustered"

cats_c  = Reference(hws, min_col=1, max_col=1, min_row=2, max_row=7)
vals_c1 = Reference(hws, min_col=2, max_col=2, min_row=1, max_row=7)
vals_c2 = Reference(hws, min_col=3, max_col=3, min_row=1, max_row=7)
col_c.add_data(vals_c1, titles_from_data=True)
col_c.add_data(vals_c2, titles_from_data=True)
col_c.set_categories(cats_c)
col_c.series[0].graphicalProperties.solidFill = "1976D2"
col_c.series[1].graphicalProperties.solidFill = "166534"

col_c.width  = 13
col_c.height = 11
ws.add_chart(col_c, "T17")

# ─── Chart D: Line – Grade Trend by Month (Grade 6-11) ───────────────────────
ws.row_dimensions[43].height = 26
section_hdr(ws, 43, 2, 28, "  MONTHLY GRADE TREND   —   How each grade's lead count changes by month")

line_d = LineChart()
line_d.title = "Monthly Leads Trend by Grade"
line_d.style = 10
line_d.y_axis.title = "Leads"
line_d.x_axis.title = "Month"

cats_d = Reference(hws, min_col=1, max_col=1, min_row=11, max_row=22)
for gi, g in enumerate(GRADES):
    vals_d = Reference(hws, min_col=2+gi, max_col=2+gi, min_row=10, max_row=22)
    line_d.add_data(vals_d, titles_from_data=True)

line_d.set_categories(cats_d)

grade_line_colors = [GRADE_COLORS[g] for g in GRADES]
for si, color in enumerate(grade_line_colors):
    if si < len(line_d.series):
        s = line_d.series[si]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 20000
        s.smooth = True

line_d.width  = 28
line_d.height = 11
ws.add_chart(line_d, "B44")

# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
wb.save(DST)
print("Dashboard 2 (Grade Interest) saved.")
print("Sheets:", wb.sheetnames)
