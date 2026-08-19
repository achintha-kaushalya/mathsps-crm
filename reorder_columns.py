#!/usr/bin/env python3
"""
Reorders columns in the "Master Leads" sheet of Master_Leads_GoogleSheets.xlsx to:
A: Phone (Normalized)
B: Raw Phone (as entered)
C: F-Code
D: Assigned Member
E: Date Added
F: Status
G: Grade
H: Comments
I: Campaign / Boost Name
J: Repeat Student?
K: Duplicate Check
L: Previous F-Code (if repeat)
M: Second Call Done
N: Second Call Notes
O: Paid
P: Grade (Final / Interested)
"""

import openpyxl

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Master_Leads_GoogleSheets.xlsx'

wb = openpyxl.load_workbook(DST)
ws = wb['Master Leads']

# The headers we want in the new order
new_headers = [
    "Phone (Normalized)",
    "Raw Phone (as entered)",
    "F-Code",
    "Assigned Member",
    "Date Added",
    "Status",
    "Grade",
    "Comments",
    "Campaign / Boost Name",
    "Repeat Student?",
    "Duplicate Check",
    "Previous F-Code (if repeat)",
    "Second Call Done",
    "Second Call Notes",
    "Paid",
    "Grade (Final / Interested)"
]

# Read existing data rows to map them properly
old_headers = [cell.value for cell in ws[1]]

# Map header name to list of values in old sheet
data_map = {h: [] for h in new_headers}
max_row = ws.max_row

if max_row >= 2:
    for r in range(2, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, len(old_headers) + 1)]
        # Map values
        for old_h, val in zip(old_headers, row_vals):
            if old_h in data_map:
                data_map[old_h].append(val)

# Delete existing rows and recreate with new column structure
ws.delete_rows(1, max_row + 1)

# Write new headers
for col_idx, h in enumerate(new_headers, 1):
    ws.cell(1, col_idx, h)

# Write data rows
if max_row >= 2:
    num_rows = max_row - 1
    for r_idx in range(num_rows):
        for col_idx, h in enumerate(new_headers, 1):
            ws.cell(r_idx + 2, col_idx, data_map[h][r_idx])

# Save workbook
wb.save(DST)
print("Reordered columns in Master Leads successfully!")
