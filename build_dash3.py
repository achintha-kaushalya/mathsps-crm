#!/usr/bin/env python3
"""
Lead CRM - Dashboard 3: Paid Dashboard
Shows paid students by grade with KPI cards and charts.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Lead_CRM_Dashboards.xlsx'
ML  = "'Master Leads'"

NAVY    = "0D2744"; BLUE   = "1565C0"; ROYAL  = "1976D2"
WHITE   = "FFFFFF"; OFF_W  = "F0F4F8"; LGRAY  = "E2E8F0"
GRAY    = "64748B"; DGRAY  = "1E293B"; GREEN  = "166534"
LGREEN  = "DCFCE7"; TEAL   = "134E4A"; LTEAL  = "CCFBF1"
AMBER   = "C2410C"; LAMBER = "FFF7ED"; RED    = "991B1B"
LRED    = "FEE2E2"; PURPLE = "581C87"; LPURPLE = "F3E8FF"
CARD_BG = "FFFFFF"

GRADE_COLORS = {
    "6": "1D4ED8", "7": "0369A1", "8": "047857",
    "9": "7C3AED", "10": "B45309", "11": "B91C1C"
}

def pf(c):   return PatternFill("solid", fgColor=c)
def ft(size=10, bold=False, color="000000", name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bd_all(color="D1D5DB"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def merge_set(ws, r1, c1, r2, c2, val=None, fill_color=None,
              font=None, align=None, numfmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if val is not None:   cell.value = val
    if font:              cell.font  = font
    if align:             cell.alignment = align
    if numfmt:            cell.number_format = numfmt
    if fill_color:
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).fill = pf(fill_color)
    return cell

def kpi_mini(ws, row, col, title, formula, sublabel, accent, numfmt=None):
    W = 5
    for r in range(row, row+3):
        for c in range(col, col+W):
            ws.cell(r, c).fill = pf(accent if r == row else CARD_BG)
    acc_s  = Side(border_style="medium", color=accent)
    gray_s = Side(border_style="thin",   color="D1D5DB")
    none_s = Side(border_style=None)
    for r in range(row, row+3):
        for c in range(col, col+W):
            ws.cell(r, c).border = Border(
                left   = acc_s  if c == col   else none_s,
                right  = gray_s if c == col+W-1 else none_s,
                top    = gray_s if r == row     else none_s,
                bottom = gray_s if r == row+2   else none_s
            )
    cell = merge_set(ws, row, col, row, col+W-1,
                     val=title.upper(), fill_color=accent,
                     font=ft(8, True, WHITE), align=al("left", "center"))
    vc = merge_set(ws, row+1, col, row+1, col+W-1,
                   val=formula, fill_color=CARD_BG,
                   font=ft(20, True, accent), align=al("left", "center"))
    if numfmt: vc.number_format = numfmt
    merge_set(ws, row+2, col, row+2, col+W-1,
              val=sublabel, fill_color=CARD_BG,
              font=ft(8, False, GRAY), align=al("left", "top"))

def section_hdr(ws, row, c1, c2, label):
    for c in range(1, c2+2): ws.cell(row, c).fill = pf(LGRAY)
    merge_set(ws, row, c1, row, c2, val=label, fill_color=LGRAY,
              font=ft(10, True, DGRAY), align=al("left", "center"))

# ═══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(DST)

HELPER = "_DB3_Data"
if HELPER in wb.sheetnames: del wb[HELPER]
hws = wb.create_sheet(HELPER); hws.sheet_state = "hidden"

DNAME = "Paid Dashboard"
if DNAME in wb.sheetnames: del wb[DNAME]
ws = wb.create_sheet(DNAME, 2)
ws.sheet_properties.tabColor = GREEN
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 85

# ── Sizing ────────────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 1.5
for c in range(2, 35): ws.column_dimensions[get_column_letter(c)].width = 3.8
for c in range(35, 70): ws.column_dimensions[get_column_letter(c)].width = 13

for r, h in {1:7, 2:52, 3:20, 4:12, 5:14, 6:38, 7:15, 8:12,
             9:14, 10:38, 11:15, 12:12, 13:26}.items():
    ws.row_dimensions[r].height = h
for r in range(14, 90): ws.row_dimensions[r].height = 14

for r in range(1, 85):
    for c in range(1, 30): ws.cell(r, c).fill = pf(OFF_W)

# ═══════════════════════════════════════════════════════════════════════════════
# TITLE
# ═══════════════════════════════════════════════════════════════════════════════
for c in range(1, 30): ws.cell(2, c).fill = pf(GREEN)
merge_set(ws, 2, 2, 2, 28, val="  PAID STUDENT DASHBOARD   |   Fee-Confirmed Enrollments",
    fill_color=GREEN, font=ft(20, True, WHITE), align=al("left", "center"))
for c in range(1, 30): ws.cell(3, c).fill = pf(TEAL)
merge_set(ws, 3, 2, 3, 28,
    val="  Shows all students who have confirmed payment  \u00b7  Grade-wise breakdown  \u00b7  Conversion rates",
    fill_color=TEAL, font=ft(9, False, "A7F3D0"), align=al("left", "center"))

# ═══════════════════════════════════════════════════════════════════════════════
# KPI CARDS
# ═══════════════════════════════════════════════════════════════════════════════
GRADES = ["6","7","8","9","10","11"]
TOTAL_LEADS = f"COUNTA({ML}!B:B)-1"
TOTAL_PAID  = f'COUNTIF({ML}!O:O,"Yes")'

kpis1 = [
    (2,  "TOTAL PAID", f"={TOTAL_PAID}",
     "All confirmed paid students", GREEN, None),
    (8,  "TOTAL LEADS", f"={TOTAL_LEADS}",
     "All leads in system", NAVY, None),
    (14, "OVERALL PAID %",
     f"=IFERROR({TOTAL_PAID}/({TOTAL_LEADS}),0)",
     "Paid ÷ Total Leads", TEAL, "0.0%"),
    (20, "INTERESTED \u2192 PAID",
     f'=IFERROR({TOTAL_PAID}/(COUNTIF({ML}!F:F,"Interested")+{TOTAL_PAID}),0)',
     "Of interested leads, how many paid", PURPLE, "0.0%"),
]
for col, title, formula, sublabel, accent, numfmt in kpis1:
    kpi_mini(ws, 5, col, title, formula, sublabel, accent, numfmt)

# KPI cards row 2: paid per grade
for gi, g in enumerate(GRADES):
    col = 2 + gi * 5 - (gi // 4) * 30   # wrap to 2 rows
    row = 9 if gi < 4 else 13
    if gi == 4: ws.row_dimensions[12].height = 10
    kpi_mini(ws, row, 2 + gi * 5 if gi < 4 else 2 + (gi-4)*5,
             f"GRADE {g} PAID",
             f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")',
             f"Paid students in Grade {g}",
             GRADE_COLORS[g], None)

# Arrange 6 grade cards: 4 in one row, 2 in next
ws.row_dimensions[9].height = 14
ws.row_dimensions[10].height = 38
ws.row_dimensions[11].height = 15
ws.row_dimensions[12].height = 10
ws.row_dimensions[13].height = 14
ws.row_dimensions[14].height = 38
ws.row_dimensions[15].height = 15
ws.row_dimensions[16].height = 12

for gi, g in enumerate(GRADES):
    if gi < 4:
        col = 2 + gi * 5
        kpi_mini(ws, 9, col, f"GRADE {g} PAID",
                 f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")',
                 f"Paid in Grade {g}",
                 GRADE_COLORS[g])
    else:
        col = 2 + (gi-4) * 5
        kpi_mini(ws, 13, col, f"GRADE {g} PAID",
                 f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")',
                 f"Paid in Grade {g}",
                 GRADE_COLORS[g])

# ═══════════════════════════════════════════════════════════════════════════════
# PAID TABLE
# ═══════════════════════════════════════════════════════════════════════════════
section_hdr(ws, 17, 2, 28, "  GRADE-WISE PAID BREAKDOWN   —   Detailed conversion per grade")

# Table header row 18
ws.row_dimensions[18].height = 22
headers_p = [("Grade", 3, NAVY), ("Total Leads", 4, BLUE),
             ("Paid", 4, GREEN), ("Not Paid", 4, AMBER),
             ("Paid %", 3, TEAL), ("Interested→Paid", 5, PURPLE)]
col = 2
for hdr_text, span, hcol in headers_p:
    merge_set(ws, 18, col, 18, col+span-1, val=hdr_text, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

# Grade rows
for gi, g in enumerate(GRADES):
    r = 19 + gi
    ws.row_dimensions[r].height = 20
    bg = LTEAL if gi % 2 == 0 else WHITE

    col = 2
    merge_set(ws, r, col, r, col+2, val=f"Grade {g}", fill_color=GRADE_COLORS[g],
              font=ft(10, True, WHITE), align=al("center", "center")); col += 3

    leads_f = f"=COUNTIF({ML}!G:G,{g})"
    merge_set(ws, r, col, r, col+3, val=leads_f, fill_color=bg,
              font=ft(10, True, DGRAY), align=al("center", "center")); col += 4

    paid_f = f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'
    merge_set(ws, r, col, r, col+3, val=paid_f, fill_color=bg,
              font=ft(10, True, GREEN), align=al("center", "center")); col += 4

    notpaid_f = f'=COUNTIF({ML}!G:G,{g})-COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'
    merge_set(ws, r, col, r, col+3, val=notpaid_f, fill_color=bg,
              font=ft(10, True, AMBER), align=al("center", "center")); col += 4

    pct_f = f'=IFERROR(COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")/COUNTIF({ML}!G:G,{g}),0)'
    c1 = merge_set(ws, r, col, r, col+2, val=pct_f, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c1.number_format = "0.0%"; col += 3

    int_paid_f = (
        f'=IFERROR(COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")/'
        f'(COUNTIFS({ML}!G:G,{g},{ML}!F:F,"Interested")+COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")),0)'
    )
    c2 = merge_set(ws, r, col, r, col+4, val=int_paid_f, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c2.number_format = "0.0%"

# Total row
r_tot = 25
ws.row_dimensions[r_tot].height = 22
for c in range(2, 27): ws.cell(r_tot, c).fill = pf(NAVY)
merge_set(ws, r_tot, 2, r_tot, 4, val="TOTAL", fill_color=NAVY,
          font=ft(10, True, WHITE), align=al("center", "center"))
merge_set(ws, r_tot, 5, r_tot, 8, val=f"={TOTAL_LEADS}", fill_color=NAVY,
          font=ft(11, True, WHITE), align=al("center", "center"))
merge_set(ws, r_tot, 9, r_tot, 12, val=f"={TOTAL_PAID}", fill_color=NAVY,
          font=ft(11, True, WHITE), align=al("center", "center"))

# ═══════════════════════════════════════════════════════════════════════════════
# HELPER DATA
# ═══════════════════════════════════════════════════════════════════════════════
hws.cell(1,1).value = "Grade"; hws.cell(1,2).value = "Leads"
hws.cell(1,3).value = "Paid";  hws.cell(1,4).value = "Not Paid"
for i, g in enumerate(GRADES):
    hws.cell(2+i, 1).value = f"Grade {g}"
    hws.cell(2+i, 2).value = f"=COUNTIF({ML}!G:G,{g})"
    hws.cell(2+i, 3).value = f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'
    hws.cell(2+i, 4).value = f'=COUNTIF({ML}!G:G,{g})-COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'

# ═══════════════════════════════════════════════════════════════════════════════
# CHARTS
# ═══════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[26].height = 26
section_hdr(ws, 26, 2, 28, "  PAID CHARTS   —   Visual breakdown of payments by grade")

# Pie: Paid by Grade
pie1 = PieChart()
pie1.title = "Paid Students by Grade"
pie1.style = 10
cats_p = Reference(hws, min_col=1, max_col=1, min_row=2, max_row=7)
vals_p = Reference(hws, min_col=3, max_col=3, min_row=1, max_row=7)
pie1.add_data(vals_p, titles_from_data=True)
pie1.set_categories(cats_p)
for idx, g in enumerate(GRADES):
    pt = DataPoint(idx=idx); pt.graphicalProperties.solidFill = GRADE_COLORS[g]
    pie1.series[0].dPt.append(pt)
pie1.width = 12; pie1.height = 11
ws.add_chart(pie1, "B27")

# Column: Leads vs Paid by Grade
col2 = BarChart()
col2.type = "col"; col2.title = "Leads vs Paid by Grade"
col2.style = 10; col2.grouping = "clustered"
col2.x_axis.title = "Grade"; col2.y_axis.title = "Count"
cats_c = Reference(hws, min_col=1, max_col=1, min_row=2, max_row=7)
v_leads = Reference(hws, min_col=2, max_col=2, min_row=1, max_row=7)
v_paid  = Reference(hws, min_col=3, max_col=3, min_row=1, max_row=7)
col2.add_data(v_leads, titles_from_data=True)
col2.add_data(v_paid,  titles_from_data=True)
col2.set_categories(cats_c)
col2.series[0].graphicalProperties.solidFill = "1976D2"
col2.series[1].graphicalProperties.solidFill = "166534"
col2.width = 14; col2.height = 11
ws.add_chart(col2, "L27")

# Doughnut: Paid vs Unpaid overall
donut3 = DoughnutChart()
donut3.title = "Overall Paid vs Unpaid"; donut3.style = 10; donut3.holeSize = 55
hws.cell(10,1).value = "Paid";   hws.cell(10,2).value = f'=COUNTIF({ML}!O:O,"Yes")'
hws.cell(11,1).value = "Unpaid"; hws.cell(11,2).value = f'=MAX(0,COUNTA({ML}!B:B)-1-COUNTIF({ML}!O:O,"Yes"))'
cats_d = Reference(hws, min_col=1, max_col=1, min_row=10, max_row=11)
vals_d = Reference(hws, min_col=2, max_col=2, min_row=9, max_row=11)
hws.cell(9,2).value = "Count"
donut3.add_data(vals_d, titles_from_data=True)
donut3.set_categories(cats_d)
pt_y = DataPoint(idx=0); pt_y.graphicalProperties.solidFill = GREEN
pt_n = DataPoint(idx=1); pt_n.graphicalProperties.solidFill = LGRAY
donut3.series[0].dPt.extend([pt_y, pt_n])
donut3.width = 10; donut3.height = 11
ws.add_chart(donut3, "T27")

wb.save(DST)
print("Dashboard 3 (Paid) saved.")
