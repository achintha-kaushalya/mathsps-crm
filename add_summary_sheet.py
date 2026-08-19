#!/usr/bin/env python3
"""
Updates the "Leads Summary" sheet with robust SUMPRODUCT + SEARCH formulas
to correctly count both numeric grades (e.g., 8) and comma-separated grade strings (e.g., "7, 8").
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Master_Leads_GoogleSheets.xlsx'

NAVY    = "0D2744"; BLUE   = "1565C0"; ROYAL  = "1976D2"; SKY = "0284C7"
WHITE   = "FFFFFF"; LGRAY  = "E2E8F0"; GRAY   = "64748B"
DGRAY   = "1E293B"; GREEN  = "166534"; LGREEN = "DCFCE7"
TEAL    = "134E4A"; LTEAL  = "CCFBF1"; AMBER  = "C2410C"
LAMBER  = "FFF7ED"; RED    = "991B1B"; LRED   = "FEE2E2"
INDIGO  = "312E81"; LINDIGO= "E0E7FF"; CARD_BG= "FFFFFF"
ORANGE  = "EA580C"

def pf(c):   return PatternFill("solid", fgColor=c)
def ft(size=10, bold=False, color="000000", name="Calibri", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_bd():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)
def med_bd():
    s = Side(border_style="medium", color=NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

# Load workbook
wb = load_workbook(DST)

# Remove existing Leads Summary if any
if "Leads Summary" in wb.sheetnames:
    del wb["Leads Summary"]

ws = wb.create_sheet("Leads Summary", 1) # Second tab
ws.sheet_properties.tabColor = ROYAL
ws.sheet_view.showGridLines = True

# Column widths
ws.column_dimensions["A"].width = 24  # Names / Campaigns
for c in ["B","C","D","E","F","G"]:
    ws.column_dimensions[c].width = 11  # Grades / Metrics
ws.column_dimensions["H"].width = 14  # Total Leads / Paid
ws.column_dimensions["I"].width = 15  # Share % / In-Progress
ws.column_dimensions["J"].width = 16  # Conv % / Actionable Rate

# Background fill
for r in range(1, 120):
    for c in range(1, 11):
        ws.cell(r, c).fill = pf("F8FAFC")

# Title Banner
ws.merge_cells("A2:J2")
title_cell = ws["A2"]
title_cell.value = "  CRM LEADS & PERFORMANCE ANALYTICS"
title_cell.fill = pf(NAVY)
title_cell.font = ft(16, True, WHITE)
title_cell.alignment = al("left", "center")
ws.row_dimensions[2].height = 35

ws.merge_cells("A3:J3")
sub_cell = ws["A3"]
sub_cell.value = "  Real-time team analytics, grade distribution, conversion metrics, unresolved pipelines, and campaign effectiveness."
sub_cell.fill = pf(ROYAL)
sub_cell.font = ft(9, False, "BFDBFE")
sub_cell.alignment = al("left", "center")
ws.row_dimensions[3].height = 20

# ─────────────────────────────────────────────────────────────────────────────
# DATE RANGE CONTROLS (Row 5)
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[5].height = 26
for c in range(1, 11):
    ws.cell(5, c).fill = pf(LINDIGO)
    ws.cell(5, c).border = thin_bd()

ws.cell(5, 1).value = "  Filter Start Date:"
ws.cell(5, 1).font = ft(10, True, INDIGO)
ws.cell(5, 1).alignment = al("left", "center")

# Start Date input field in B5
ws.cell(5, 2).fill = pf(WHITE)
ws.cell(5, 2).border = med_bd()
ws.cell(5, 2).font = ft(10, True, DGRAY)
ws.cell(5, 2).number_format = "YYYY-MM-DD"
ws.cell(5, 2).alignment = al("center", "center")

ws.cell(5, 3).value = "End Date:"
ws.cell(5, 3).font = ft(10, True, INDIGO)
ws.cell(5, 3).alignment = al("right", "center")

# End Date input field in D5
ws.cell(5, 4).fill = pf(WHITE)
ws.cell(5, 4).border = med_bd()
ws.cell(5, 4).font = ft(10, True, DGRAY)
ws.cell(5, 4).number_format = "YYYY-MM-DD"
ws.cell(5, 4).alignment = al("center", "center")

ws.merge_cells("E5:J5")
info_cell = ws["E5"]
info_cell.value = "Leave dates blank to view All-Time data. Format: YYYY-MM-DD (e.g. 2026-07-01)"
info_cell.font = ft(9, False, GRAY, italic=True)
info_cell.alignment = al("left", "center")

# ─────────────────────────────────────────────────────────────────────────────
# TABLE 1: ALL LEADS BY MEMBER & GRADE
# ─────────────────────────────────────────────────────────────────────────────
start_r1 = 7
ws.cell(start_r1, 1).value = "TABLE 1: ALL LEADS BY MEMBER & GRADE"
ws.cell(start_r1, 1).font = ft(11, True, NAVY)
ws.row_dimensions[start_r1].height = 24

h_row1 = start_r1 + 1
ws.row_dimensions[h_row1].height = 24
headers1 = ["Assigned Member", "Grade 6", "Grade 7", "Grade 8", "Grade 9", "Grade 10", "Grade 11", "Total Leads", "Share %", ""]
for c_idx, h_text in enumerate(headers1, 1):
    if c_idx == 10: continue
    cell = ws.cell(h_row1, c_idx)
    cell.value = h_text
    cell.fill = pf(NAVY if c_idx in [1, 8, 9] else BLUE)
    cell.font = ft(9, True, WHITE)
    cell.alignment = al("center", "center")
    cell.border = med_bd()

max_members = 20
curr_r = h_row1 + 1
ML = "'Master Leads'"

for i in range(max_members):
    ws.row_dimensions[curr_r].height = 20
    bg = "FFFFFF" if i % 2 == 0 else "F1F5F9"
    
    cell_a = ws.cell(curr_r, 1)
    cell_a.value = f'=IF(INDEX(Members!$A:$A,{i+2})="","",INDEX(Members!$A:$A,{i+2}))'
    cell_a.font = ft(10, True, DGRAY)
    cell_a.fill = pf(bg)
    cell_a.border = thin_bd()
    cell_a.alignment = al("left", "center")
    
    for g_idx, grade in enumerate([6, 7, 8, 9, 10, 11], 2):
        cell_g = ws.cell(curr_r, g_idx)
        # SUMPRODUCT + SEARCH logic to support both number (8) and string ("7, 8")
        cell_g.value = (
            f'=IF($A{curr_r}="","",SUMPRODUCT('
            f'({ML}!$D:$D=$A{curr_r}) * '
            f'(ISNUMBER(SEARCH("{grade}", {ML}!$G:$G))) * '
            f'({ML}!$E:$E>=IF(ISBLANK($B$5),0,$B$5)) * '
            f'({ML}!$E:$E<=IF(ISBLANK($D$5),99999,$D$5))'
            f'))'
        )
        cell_g.font = ft(10, False, DGRAY)
        cell_g.fill = pf(bg)
        cell_g.border = thin_bd()
        cell_g.alignment = al("center", "center")
        
    # H - Total Leads (calculated directly from Master Leads to count actual phone leads/rows)
    cell_tot = ws.cell(curr_r, 8)
    cell_tot.value = (
        f'=IF($A{curr_r}="","",COUNTIFS('
        f'\'Master Leads\'!$D:$D,$A{curr_r},'
        f'\'Master Leads\'!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),'
        f'\'Master Leads\'!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5)'
        f'))'
    )
    cell_tot.font = ft(10, True, BLUE)
    cell_tot.fill = pf(bg)
    cell_tot.border = thin_bd()
    cell_tot.alignment = al("center", "center")

    # I - Lead Share % (vs Grand Total)
    cell_sh = ws.cell(curr_r, 9)
    # The Grand Total will be on row tot_r1 (defined below)
    tot_row_num = h_row1 + max_members + 1
    cell_sh.value = f'=IF(OR($A{curr_r}="",$H${tot_row_num}=0),"",H{curr_r}/$H${tot_row_num})'
    cell_sh.font = ft(10, False, DGRAY)
    cell_sh.fill = pf(bg)
    cell_sh.border = thin_bd()
    cell_sh.number_format = "0.0%"
    cell_sh.alignment = al("center", "center")
    
    curr_r += 1

# Total Row for Table 1
tot_r1 = curr_r
ws.row_dimensions[tot_r1].height = 22
ws.cell(tot_r1, 1).value = "TOTAL"
ws.cell(tot_r1, 1).font = ft(10, True, WHITE)
ws.cell(tot_r1, 1).fill = pf(NAVY)
ws.cell(tot_r1, 1).alignment = al("center", "center")
ws.cell(tot_r1, 1).border = thin_bd()

for c_idx in range(2, 8):
    grade_val = [6, 7, 8, 9, 10, 11][c_idx-2]
    cell = ws.cell(tot_r1, c_idx)
    cell.value = (
        f'=SUMPRODUCT('
        f'(ISNUMBER(SEARCH("{grade_val}", {ML}!$G:$G))) * '
        f'({ML}!$E:$E>=IF(ISBLANK($B$5),0,$B$5)) * '
        f'({ML}!$E:$E<=IF(ISBLANK($D$5),99999,$D$5))'
        f')'
    )
    cell.font = ft(10, True, WHITE)
    cell.fill = pf(NAVY)
    cell.alignment = al("center", "center")
    cell.border = thin_bd()

# Grand Total Leads in Table 1 (Count actual leads / rows)
ws.cell(tot_r1, 8).value = (
    f'=COUNTIFS('
    f'\'Master Leads\'!$D:$D,"<>",'
    f'\'Master Leads\'!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),'
    f'\'Master Leads\'!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5)'
    f')'
)
ws.cell(tot_r1, 8).font = ft(10, True, WHITE)
ws.cell(tot_r1, 8).fill = pf(NAVY)
ws.cell(tot_r1, 8).alignment = al("center", "center")
ws.cell(tot_r1, 8).border = thin_bd()

ws.cell(tot_r1, 9).value = 1.0
ws.cell(tot_r1, 9).font = ft(10, True, WHITE)
ws.cell(tot_r1, 9).fill = pf(NAVY)
ws.cell(tot_r1, 9).number_format = "0.0%"
ws.cell(tot_r1, 9).alignment = al("center", "center")
ws.cell(tot_r1, 9).border = thin_bd()


# ─────────────────────────────────────────────────────────────────────────────
# TABLE 2: PAID LEADS BY MEMBER & GRADE
# ─────────────────────────────────────────────────────────────────────────────
start_r2 = tot_r1 + 3
ws.cell(start_r2, 1).value = "TABLE 2: PAID LEADS BY MEMBER & GRADE"
ws.cell(start_r2, 1).font = ft(11, True, GREEN)
ws.row_dimensions[start_r2].height = 24

h_row2 = start_r2 + 1
ws.row_dimensions[h_row2].height = 24
headers2 = ["Assigned Member", "Grade 6", "Grade 7", "Grade 8", "Grade 9", "Grade 10", "Grade 11", "Total Paid", "Paid Share %", "Conversion %"]
for c_idx, h_text in enumerate(headers2, 1):
    cell = ws.cell(h_row2, c_idx)
    cell.value = h_text
    cell.fill = pf(GREEN if c_idx in [1, 8, 9, 10] else TEAL)
    cell.font = ft(9, True, WHITE)
    cell.alignment = al("center", "center")
    cell.border = med_bd()

curr_r = h_row2 + 1
tot_row2_num = h_row2 + max_members + 1

for i in range(max_members):
    ws.row_dimensions[curr_r].height = 20
    bg = "FFFFFF" if i % 2 == 0 else "F1F5F9"
    
    cell_a = ws.cell(curr_r, 1)
    cell_a.value = f'=IF(INDEX(Members!$A:$A,{i+2})="","",INDEX(Members!$A:$A,{i+2}))'
    cell_a.font = ft(10, True, DGRAY)
    cell_a.fill = pf(bg)
    cell_a.border = thin_bd()
    cell_a.alignment = al("left", "center")
    
    for g_idx, grade in enumerate([6, 7, 8, 9, 10, 11], 2):
        cell_g = ws.cell(curr_r, g_idx)
        # SUMPRODUCT + SEARCH logic on Column P (Grade Final/Interested) + Paid="Yes"
        cell_g.value = (
            f'=IF($A{curr_r}="","",SUMPRODUCT('
            f'({ML}!$D:$D=$A{curr_r}) * '
            f'(ISNUMBER(SEARCH("{grade}", {ML}!$P:$P))) * '
            f'({ML}!$O:$O="Yes") * '
            f'({ML}!$E:$E>=IF(ISBLANK($B$5),0,$B$5)) * '
            f'({ML}!$E:$E<=IF(ISBLANK($D$5),99999,$D$5))'
            f'))'
        )
        cell_g.font = ft(10, False, DGRAY)
        cell_g.fill = pf(bg)
        cell_g.border = thin_bd()
        cell_g.alignment = al("center", "center")
        
    cell_tot = ws.cell(curr_r, 8)
    cell_tot.value = f'=IF($A{curr_r}="","",SUM(B{curr_r}:G{curr_r}))'
    cell_tot.font = ft(10, True, GREEN)
    cell_tot.fill = pf(bg)
    cell_tot.border = thin_bd()
    cell_tot.alignment = al("center", "center")

    cell_ps = ws.cell(curr_r, 9)
    cell_ps.value = f'=IF(OR($A{curr_r}="",$H${tot_row2_num}=0),"",H{curr_r}/$H${tot_row2_num})'
    cell_ps.font = ft(10, False, DGRAY)
    cell_ps.fill = pf(bg)
    cell_ps.border = thin_bd()
    cell_ps.number_format = "0.0%"
    cell_ps.alignment = al("center", "center")

    t1_r = curr_r - (h_row2 - h_row1)
    cell_cv = ws.cell(curr_r, 10)
    cell_cv.value = f'=IF(OR($A{curr_r}="",H{t1_r}=0),"",H{curr_r}/H{t1_r})'
    cell_cv.font = ft(10, True, TEAL)
    cell_cv.fill = pf(bg)
    cell_cv.border = thin_bd()
    cell_cv.number_format = "0.0%"
    cell_cv.alignment = al("center", "center")
    
    curr_r += 1

tot_r2 = curr_r
ws.row_dimensions[tot_r2].height = 22
ws.cell(tot_r2, 1).value = "TOTAL PAID"
ws.cell(tot_r2, 1).font = ft(10, True, WHITE)
ws.cell(tot_r2, 1).fill = pf(GREEN)
ws.cell(tot_r2, 1).alignment = al("center", "center")
ws.cell(tot_r2, 1).border = thin_bd()

for c_idx in range(2, 8):
    grade_val = [6, 7, 8, 9, 10, 11][c_idx-2]
    cell = ws.cell(tot_r2, c_idx)
    cell.value = (
        f'=SUMPRODUCT('
        f'(ISNUMBER(SEARCH("{grade_val}", {ML}!$P:$P))) * '
        f'({ML}!$O:$O="Yes") * '
        f'({ML}!$E:$E>=IF(ISBLANK($B$5),0,$B$5)) * '
        f'({ML}!$E:$E<=IF(ISBLANK($D$5),99999,$D$5))'
        f')'
    )
    cell.font = ft(10, True, WHITE)
    cell.fill = pf(GREEN)
    cell.alignment = al("center", "center")
    cell.border = thin_bd()

ws.cell(tot_r2, 8).value = f'=SUM(B{tot_r2}:G{tot_r2})'
ws.cell(tot_r2, 8).font = ft(10, True, WHITE)
ws.cell(tot_r2, 8).fill = pf(GREEN)
ws.cell(tot_r2, 8).alignment = al("center", "center")
ws.cell(tot_r2, 8).border = thin_bd()

ws.cell(tot_r2, 9).value = 1.0
ws.cell(tot_r2, 9).font = ft(10, True, WHITE)
ws.cell(tot_r2, 9).fill = pf(GREEN)
ws.cell(tot_r2, 9).number_format = "0.0%"
ws.cell(tot_r2, 9).alignment = al("center", "center")
ws.cell(tot_r2, 9).border = thin_bd()

ws.cell(tot_r2, 10).value = f'=IF(H{tot_r1}=0,"",H{tot_r2}/H{tot_r1})'
ws.cell(tot_r2, 10).font = ft(10, True, WHITE)
ws.cell(tot_r2, 10).fill = pf(GREEN)
ws.cell(tot_r2, 10).number_format = "0.0%"
ws.cell(tot_r2, 10).alignment = al("center", "center")
ws.cell(tot_r2, 10).border = thin_bd()


# ─────────────────────────────────────────────────────────────────────────────
# TABLE 3: UNRESOLVED / PIPELINE STATUS BY MEMBER
# ─────────────────────────────────────────────────────────────────────────────
start_r3 = tot_r2 + 3
ws.cell(start_r3, 1).value = "TABLE 3: UNRESOLVED LEADS PIPELINE BY MEMBER"
ws.cell(start_r3, 1).font = ft(11, True, AMBER)
ws.row_dimensions[start_r3].height = 24

h_row3 = start_r3 + 1
ws.row_dimensions[h_row3].height = 24
headers3 = ["Assigned Member", "New Leads", "Contacted", "No Answer", "Interested", "Other Pending", "Total Unresolved", "Unresolved %", "Actionable Rate", ""]
for c_idx, h_text in enumerate(headers3, 1):
    if c_idx == 10: continue
    cell = ws.cell(h_row3, c_idx)
    cell.value = h_text
    cell.fill = pf(AMBER if c_idx in [1, 7, 8, 9] else ORANGE)
    cell.font = ft(9, True, WHITE)
    cell.alignment = al("center", "center")
    cell.border = med_bd()

curr_r = h_row3 + 1
tot_row3_num = h_row3 + max_members + 1

for i in range(max_members):
    ws.row_dimensions[curr_r].height = 20
    bg = "FFFFFF" if i % 2 == 0 else "F1F5F9"
    
    cell_a = ws.cell(curr_r, 1)
    cell_a.value = f'=IF(INDEX(Members!$A:$A,{i+2})="","",INDEX(Members!$A:$A,{i+2}))'
    cell_a.font = ft(10, True, DGRAY)
    cell_a.fill = pf(bg)
    cell_a.border = thin_bd()
    cell_a.alignment = al("left", "center")
    
    ws.cell(curr_r, 2).value = f'=IF($A{curr_r}="","",COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$F:$F,"New",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5)))'
    ws.cell(curr_r, 3).value = f'=IF($A{curr_r}="","",COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$F:$F,"Contacted",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5)))'
    ws.cell(curr_r, 4).value = f'=IF($A{curr_r}="","",COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$F:$F,"No Answer",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5)))'
    ws.cell(curr_r, 5).value = f'=IF($A{curr_r}="","",COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$F:$F,"Interested",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5)))'
    ws.cell(curr_r, 6).value = f'=IF($A{curr_r}="","",COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))-COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$F:$F,"Converted",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))-COUNTIFS({ML}!$D:$D,$A{curr_r},{ML}!$F:$F,"Not Interested",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))-B{curr_r}-C{curr_r}-D{curr_r}-E{curr_r})'
    
    for col_c in [2, 3, 4, 5, 6]:
        ws.cell(curr_r, col_c).font = ft(10, False, DGRAY)
        ws.cell(curr_r, col_c).fill = pf(bg)
        ws.cell(curr_r, col_c).border = thin_bd()
        ws.cell(curr_r, col_c).alignment = al("center", "center")

    ws.cell(curr_r, 7).value = f'=IF($A{curr_r}="","",SUM(B{curr_r}:F{curr_r}))'
    ws.cell(curr_r, 7).font = ft(10, True, AMBER)
    ws.cell(curr_r, 7).fill = pf(bg)
    ws.cell(curr_r, 7).border = thin_bd()
    ws.cell(curr_r, 7).alignment = al("center", "center")

    t1_r = curr_r - (h_row3 - h_row1)
    ws.cell(curr_r, 8).value = f'=IF(OR($A{curr_r}="",H{t1_r}=0),"",G{curr_r}/H{t1_r})'
    ws.cell(curr_r, 8).font = ft(10, False, DGRAY)
    ws.cell(curr_r, 8).fill = pf(bg)
    ws.cell(curr_r, 8).border = thin_bd()
    ws.cell(curr_r, 8).number_format = "0.0%"
    ws.cell(curr_r, 8).alignment = al("center", "center")

    ws.cell(curr_r, 9).value = f'=IF(OR($A{curr_r}="",H{t1_r}=0),"",(B{curr_r}+E{curr_r}+F{curr_r})/H{t1_r})'
    ws.cell(curr_r, 9).font = ft(10, True, ORANGE)
    ws.cell(curr_r, 9).fill = pf(bg)
    ws.cell(curr_r, 9).border = thin_bd()
    ws.cell(curr_r, 9).number_format = "0.0%"
    ws.cell(curr_r, 9).alignment = al("center", "center")

    curr_r += 1

tot_r3 = curr_r
ws.row_dimensions[tot_r3].height = 22
ws.cell(tot_r3, 1).value = "TOTAL UNRESOLVED"
ws.cell(tot_r3, 1).font = ft(10, True, WHITE)
ws.cell(tot_r3, 1).fill = pf(AMBER)
ws.cell(tot_r3, 1).alignment = al("center", "center")
ws.cell(tot_r3, 1).border = thin_bd()

for c_idx in range(2, 7):
    st_val = ["New", "Contacted", "No Answer", "Interested"][c_idx-2] if c_idx < 6 else "Other"
    cell = ws.cell(tot_r3, c_idx)
    if c_idx < 6:
        cell.value = f'=COUNTIFS({ML}!$F:$F,"{st_val}",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))'
    else:
        cell.value = f'=H{tot_r3}-COUNTIFS({ML}!$F:$F,"Converted",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))-COUNTIFS({ML}!$F:$F,"Not Interested",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))-SUM(B{tot_r3}:E{tot_r3})'
    cell.font = ft(10, True, WHITE)
    cell.fill = pf(AMBER)
    cell.alignment = al("center", "center")
    cell.border = thin_bd()

ws.cell(tot_r3, 7).value = f'=SUM(B{tot_r3}:F{tot_r3})'
ws.cell(tot_r3, 7).font = ft(10, True, WHITE)
ws.cell(tot_r3, 7).fill = pf(AMBER)
ws.cell(tot_r3, 7).alignment = al("center", "center")
ws.cell(tot_r3, 7).border = thin_bd()

ws.cell(tot_r3, 8).value = f'=IF(H{tot_r1}=0,"",G{tot_r3}/H{tot_r1})'
ws.cell(tot_r3, 8).font = ft(10, True, WHITE)
ws.cell(tot_r3, 8).fill = pf(AMBER)
ws.cell(tot_r3, 8).number_format = "0.0%"
ws.cell(tot_r3, 8).alignment = al("center", "center")
ws.cell(tot_r3, 8).border = thin_bd()

ws.cell(tot_r3, 9).value = f'=IF(H{tot_r1}=0,"",(B{tot_r3}+E{tot_r3}+F{tot_r3})/H{tot_r1})'
ws.cell(tot_r3, 9).font = ft(10, True, WHITE)
ws.cell(tot_r3, 9).fill = pf(AMBER)
ws.cell(tot_r3, 9).number_format = "0.0%"
ws.cell(tot_r3, 9).alignment = al("center", "center")
ws.cell(tot_r3, 9).border = thin_bd()


# ─────────────────────────────────────────────────────────────────────────────
# TABLE 4: CAMPAIGN PERFORMANCE MATRIX
# ─────────────────────────────────────────────────────────────────────────────
start_r4 = tot_r3 + 3
ws.cell(start_r4, 1).value = "TABLE 4: CAMPAIGN ROI & PERFORMANCE MATRIX"
ws.cell(start_r4, 1).font = ft(11, True, INDIGO)
ws.row_dimensions[start_r4].height = 24

h_row4 = start_r4 + 1
ws.row_dimensions[h_row4].height = 24
headers4 = ["Campaign / Boost", "Total Leads", "Converted", "Paid Students", "Not Interested", "Active Pipeline", "Conversion %", "Paid Share %", "", ""]
for c_idx, h_text in enumerate(headers4, 1):
    if c_idx in [9, 10]: continue
    cell = ws.cell(h_row4, c_idx)
    cell.value = h_text
    cell.fill = pf(INDIGO if c_idx in [1, 2, 7, 8] else SKY)
    cell.font = ft(9, True, WHITE)
    cell.alignment = al("center", "center")
    cell.border = med_bd()

campaigns = ["B1", "B2", "B3", "B4", "B5", "B6", "B7", "B8", "B9", "B10"]
tot_row4_num = h_row4 + len(campaigns) + 1

for ci, camp in enumerate(campaigns):
    r_idx = h_row4 + 1 + ci
    ws.row_dimensions[r_idx].height = 20
    bg = "FFFFFF" if ci % 2 == 0 else "F1F5F9"
    
    cell_a = ws.cell(r_idx, 1)
    cell_a.value = camp
    cell_a.font = ft(10, True, DGRAY)
    cell_a.fill = pf(bg)
    cell_a.border = thin_bd()
    cell_a.alignment = al("center", "center")
    
    ws.cell(r_idx, 2).value = f'=COUNTIFS({ML}!$I:$I,"{camp}",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))'
    ws.cell(r_idx, 3).value = f'=COUNTIFS({ML}!$I:$I,"{camp}",{ML}!$F:$F,"Converted",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))'
    ws.cell(r_idx, 4).value = f'=COUNTIFS({ML}!$I:$I,"{camp}",{ML}!$O:$O,"Yes",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))'
    ws.cell(r_idx, 5).value = f'=COUNTIFS({ML}!$I:$I,"{camp}",{ML}!$F:$F,"Not Interested",{ML}!$E:$E,">="&IF(ISBLANK($B$5),0,$B$5),{ML}!$E:$E,"<="&IF(ISBLANK($D$5),99999,$D$5))'
    ws.cell(r_idx, 6).value = f'=B{r_idx}-C{r_idx}-E{r_idx}'
    
    for col_c in [2, 3, 4, 5, 6]:
        ws.cell(r_idx, col_c).font = ft(10, False, DGRAY)
        ws.cell(r_idx, col_c).fill = pf(bg)
        ws.cell(r_idx, col_c).border = thin_bd()
        ws.cell(r_idx, col_c).alignment = al("center", "center")

    ws.cell(r_idx, 7).value = f'=IF(B{r_idx}=0,"",C{r_idx}/B{r_idx})'
    ws.cell(r_idx, 7).font = ft(10, True, TEAL)
    ws.cell(r_idx, 7).fill = pf(bg)
    ws.cell(r_idx, 7).border = thin_bd()
    ws.cell(r_idx, 7).number_format = "0.0%"
    ws.cell(r_idx, 7).alignment = al("center", "center")

    ws.cell(r_idx, 8).value = f'=IF(OR(B{r_idx}=0,$D${tot_row4_num}=0),"",D{r_idx}/$D${tot_row4_num})'
    ws.cell(r_idx, 8).font = ft(10, False, DGRAY)
    ws.cell(r_idx, 8).fill = pf(bg)
    ws.cell(r_idx, 8).border = thin_bd()
    ws.cell(r_idx, 8).number_format = "0.0%"
    ws.cell(r_idx, 8).alignment = al("center", "center")

tot_r4 = h_row4 + len(campaigns) + 1
ws.row_dimensions[tot_r4].height = 22
ws.cell(tot_r4, 1).value = "TOTAL CAMPAIGN"
ws.cell(tot_r4, 1).font = ft(10, True, WHITE)
ws.cell(tot_r4, 1).fill = pf(INDIGO)
ws.cell(tot_r4, 1).alignment = al("center", "center")
ws.cell(tot_r4, 1).border = thin_bd()

ws.cell(tot_r4, 2).value = f'=SUM(B{h_row4+1}:B{tot_r4-1})'
ws.cell(tot_r4, 3).value = f'=SUM(C{h_row4+1}:C{tot_r4-1})'
ws.cell(tot_r4, 4).value = f'=SUM(D{h_row4+1}:D{tot_r4-1})'
ws.cell(tot_r4, 5).value = f'=SUM(E{h_row4+1}:E{tot_r4-1})'
ws.cell(tot_r4, 6).value = f'=SUM(F{h_row4+1}:F{tot_r4-1})'

for c_idx in [2, 3, 4, 5, 6]:
    cell = ws.cell(tot_r4, c_idx)
    cell.font = ft(10, True, WHITE)
    cell.fill = pf(INDIGO)
    cell.alignment = al("center", "center")
    cell.border = thin_bd()

ws.cell(tot_r4, 7).value = f'=IF(B{tot_r4}=0,"",C{tot_r4}/B{tot_r4})'
ws.cell(tot_r4, 7).font = ft(10, True, WHITE)
ws.cell(tot_r4, 7).fill = pf(INDIGO)
ws.cell(tot_r4, 7).number_format = "0.0%"
ws.cell(tot_r4, 7).alignment = al("center", "center")
ws.cell(tot_r4, 7).border = thin_bd()

ws.cell(tot_r4, 8).value = 1.0
ws.cell(tot_r4, 8).font = ft(10, True, WHITE)
ws.cell(tot_r4, 8).fill = pf(INDIGO)
ws.cell(tot_r4, 8).number_format = "0.0%"
ws.cell(tot_r4, 8).alignment = al("center", "center")
ws.cell(tot_r4, 8).border = thin_bd()

# Reorder sheets to keep it clean
desired_order = ["Master Leads", "Leads Summary", "Members", "Instructions"]
for pos, name in enumerate(desired_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=pos - wb.index(wb[name]))

# Save workbook
wb.save(DST)
print("Leads Summary updated successfully with SUMPRODUCT + SEARCH formulas!")
