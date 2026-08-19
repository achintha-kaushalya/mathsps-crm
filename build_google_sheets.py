#!/usr/bin/env python3
"""
Build a clean Google-Sheets-compatible .xlsx for upload.
Sheet: Master Leads  (with existing data preserved)
Sheet: Members       (with existing members)
All formulas written in a Google Sheets-compatible style.
"""

import shutil
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Border, Side,
                              Alignment, Protection)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Master_Leads_GoogleSheets.xlsx'

# ── Existing data (from your modified workbook) ───────────────────────────────
MEMBERS = ['maheesha','menik','kavinda','umesh','Avishka','Wihanga','muthumali','shamali']

# Raw existing rows (col A-P), dates kept as strings / serial floats to convert
EXISTING_DATA = [
    # A=NormPhone,B=RawPhone,C=FCode,D=Member,E=Date,F=Status,G=Grade,
    # H=Comments,I=2ndCallDone,J=2ndCallNotes,L=Campaign,M=Repeat,N=PrevFCode,O=Paid,P=FinalGrade
    ['778625101','+94 77 862 5101','F00001','maheesha','2026-07-01','Converted','6','','No','','','B1','No','','No','7, 9'],
    ['766820157','0766820157',     'F00002','umesh',   '2026-07-02','Not Interested','','off','No','','','B1','No','','',''],
    ['712197438','+94 71 219 7438','F00003','kavinda', '2026-07-03','Contacted','11','','Yes','details ewanna','','','No','','Yes','11'],
    ['716030661','716030661',      'F00004','umesh',   '2026-07-04','Contacted','8, 9','send details','','','','','No','','','9'],
    ['781833131','781833131',      'F00005','umesh',   '2026-07-04','Converted','9','','','','','','No','','',''],
    ['716030661','716030661',      'F00004','',        '2026-07-04','New','7','','','','','','Yes - see Previous F-Code','F00004','Yes','7'],
    ['752945180','752945180',      'F00006','kavinda', '2026-07-04','No Answer','','','No','','','B1','No','','',''],
    ['752945180','94752945180',    'F00006','muthumali','2026-07-04','Contacted','8','','','','','B1','Yes - see Previous F-Code','F00006','',''],
    ['778625101','778625101',      'F00001','umesh',   '2026-07-04','Contacted','10','','','','','B2','Yes - see Previous F-Code','F00001','',''],
    ['789632546','789632546',      'F00007','',        '2026-07-04','','','','','','','','No','','',''],
]

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY    = "0D2744"; BLUE   = "1565C0"; ROYAL  = "1976D2"
WHITE   = "FFFFFF"; LGRAY  = "E2E8F0"; GRAY   = "64748B"
DGRAY   = "1E293B"; GREEN  = "166534"; LGREEN = "DCFCE7"
TEAL    = "134E4A"; LTEAL  = "CCFBF1"; AMBER  = "C2410C"
LAMBER  = "FFF7ED"; RED    = "991B1B"; LRED   = "FEE2E2"
YELLOW  = "FEF08A"; LBLUE  = "DBEAFE"; LORANGE= "FED7AA"
DPURPLE = "4C1D95"

def pf(c):   return PatternFill("solid", fgColor=c)
def ft(size=10, bold=False, color="000000", name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_bd():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)
def med_bd():
    s = Side(border_style="medium", color=NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

# ══════════════════════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — Master Leads
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Master Leads"
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = True
ws.freeze_panes = "A2"

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {
    'A': 15,  # Phone (Normalized) - hidden / formula
    'B': 18,  # Raw Phone
    'C': 10,  # F-Code
    'D': 14,  # Assigned Member
    'E': 13,  # Date Added
    'F': 16,  # Status
    'G': 8,   # Grade
    'H': 25,  # Comments
    'I': 14,  # Second Call Done
    'J': 25,  # Second Call Notes
    'K': 18,  # Duplicate Check
    'L': 18,  # Campaign / Boost Name
    'M': 22,  # Repeat Student?
    'N': 18,  # Previous F-Code
    'O': 8,   # Paid
    'P': 14,  # Grade (Final)
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Hide col A (formula column, auto-managed)
ws.column_dimensions['A'].hidden = True

# ── Header row ────────────────────────────────────────────────────────────────
headers = [
    ('A', 'Phone (Normalized)',         NAVY),
    ('B', 'Raw Phone (as entered)',      NAVY),
    ('C', 'F-Code',                     NAVY),
    ('D', 'Assigned Member',            ROYAL),
    ('E', 'Date Added',                 ROYAL),
    ('F', 'Status',                     ROYAL),
    ('G', 'Grade',                      ROYAL),
    ('H', 'Comments',                   GRAY),
    ('I', 'Second Call Done',           TEAL),
    ('J', 'Second Call Notes',          TEAL),
    ('K', 'Duplicate Check',            RED),
    ('L', 'Campaign / Boost Name',      AMBER),
    ('M', 'Repeat Student?',            DPURPLE),
    ('N', 'Previous F-Code (if repeat)',DPURPLE),
    ('O', 'Paid',                       GREEN),
    ('P', 'Grade (Final / Interested)', BLUE),
]

ws.row_dimensions[1].height = 28
for col_letter, header_text, bg_color in headers:
    col_idx = ord(col_letter) - ord('A') + 1
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header_text
    cell.fill  = pf(bg_color)
    cell.font  = ft(10, True, WHITE)
    cell.alignment = al("center", "center")
    cell.border = med_bd()

# ── Data rows (formulas for A, K, M, N + hardcoded data for B-P) ─────────────
STATUS_OPTIONS  = ["New","Contacted","Interested","Converted",
                   "No Answer","Not Interested","Follow-up","Second Call Pending"]
GRADE_OPTIONS   = ["6","7","8","9","10","11"]
PAID_OPTIONS    = ["Yes","No"]
SECONDCALL_OPT  = ["Yes","No"]
YES_NO          = ["Yes","No"]

LAST_DATA_ROW   = 1 + len(EXISTING_DATA)   # row index of last data row

for i, row_data in enumerate(EXISTING_DATA, start=2):
    (norm_phone, raw_phone, fcode, member, date_str, status, grade,
     comments, second_call, second_notes, _dup_check, campaign,
     repeat, prev_fcode, paid, final_grade) = row_data

    ws.row_dimensions[i].height = 20

    # A – Phone Normalized (Google Sheets REGEXREPLACE formula)
    ws.cell(i, 1).value = (
        f'=IF(B{i}="","",RIGHT(REGEXREPLACE(B{i}&"","[^0-9]",""),9))'
    )
    ws.cell(i, 1).fill = pf(LGRAY)
    ws.cell(i, 1).font = ft(9, False, GRAY)
    ws.cell(i, 1).border = thin_bd()

    # B – Raw Phone
    ws.cell(i, 2).value = raw_phone
    ws.cell(i, 2).font  = ft(10)
    ws.cell(i, 2).border = thin_bd()

    # C – F-Code
    ws.cell(i, 3).value = fcode
    ws.cell(i, 3).font  = ft(10, True, NAVY)
    ws.cell(i, 3).alignment = al("center")
    ws.cell(i, 3).border = thin_bd()

    # D – Assigned Member
    ws.cell(i, 4).value = member
    ws.cell(i, 4).font  = ft(10)
    ws.cell(i, 4).border = thin_bd()

    # E – Date Added
    if date_str and '.' in date_str and len(date_str) > 10:
        # Excel serial number - convert to date
        from openpyxl.utils.datetime import from_excel
        try:
            actual_date = from_excel(float(date_str))
            ws.cell(i, 5).value = actual_date
        except:
            ws.cell(i, 5).value = date_str
    else:
        ws.cell(i, 5).value = date_str
    ws.cell(i, 5).number_format = "YYYY-MM-DD"
    ws.cell(i, 5).font = ft(10)
    ws.cell(i, 5).border = thin_bd()

    # F – Status
    ws.cell(i, 6).value = status
    ws.cell(i, 6).font  = ft(10)
    ws.cell(i, 6).border = thin_bd()

    # G – Grade
    ws.cell(i, 7).value = grade
    ws.cell(i, 7).font  = ft(10)
    ws.cell(i, 7).alignment = al("center")
    ws.cell(i, 7).border = thin_bd()

    # H – Comments
    ws.cell(i, 8).value = comments
    ws.cell(i, 8).font  = ft(10)
    ws.cell(i, 8).border = thin_bd()

    # I – Second Call Done
    ws.cell(i, 9).value = second_call
    ws.cell(i, 9).font  = ft(10)
    ws.cell(i, 9).alignment = al("center")
    ws.cell(i, 9).border = thin_bd()

    # J – Second Call Notes
    ws.cell(i, 10).value = second_notes
    ws.cell(i, 10).font  = ft(10)
    ws.cell(i, 10).border = thin_bd()

    # K – Duplicate Check (formula)
    ws.cell(i, 11).value = (
        f'=IF(COUNTIFS($A$2:A{i},A{i},$L$2:L{i},L{i})>1,"DUPLICATE - CHECK","")'
    )
    ws.cell(i, 11).font = ft(10, True, RED)
    ws.cell(i, 11).alignment = al("center")
    ws.cell(i, 11).border = thin_bd()

    # L – Campaign
    ws.cell(i, 12).value = campaign
    ws.cell(i, 12).font  = ft(10)
    ws.cell(i, 12).alignment = al("center")
    ws.cell(i, 12).border = thin_bd()

    # M – Repeat Student (formula)
    if i <= 2:
        ws.cell(i, 13).value = "No"
    else:
        ws.cell(i, 13).value = (
            f'=IF(COUNTIF($A$2:A{i-1},A{i})>0,"Yes - see Previous F-Code","No")'
        )
    ws.cell(i, 13).font = ft(10)
    ws.cell(i, 13).alignment = al("center")
    ws.cell(i, 13).border = thin_bd()

    # N – Previous F-Code (formula)
    if i <= 2:
        ws.cell(i, 14).value = ""
    else:
        ws.cell(i, 14).value = (
            f'=IF(COUNTIF($A$2:A{i-1},A{i})>0,'
            f'INDEX($C$2:C{i-1},MATCH(A{i},$A$2:A{i-1},0)),"")'
        )
    ws.cell(i, 14).font = ft(10)
    ws.cell(i, 14).alignment = al("center")
    ws.cell(i, 14).border = thin_bd()

    # O – Paid
    ws.cell(i, 15).value = paid
    ws.cell(i, 15).font  = ft(10, True, GREEN if paid == "Yes" else RED)
    ws.cell(i, 15).alignment = al("center")
    ws.cell(i, 15).border = thin_bd()

    # P – Final Grade
    ws.cell(i, 16).value = final_grade
    ws.cell(i, 16).font  = ft(10)
    ws.cell(i, 16).alignment = al("center")
    ws.cell(i, 16).border = thin_bd()

# ── Data Validations (dropdown lists) ─────────────────────────────────────────
# F – Status dropdown
dv_status = DataValidation(
    type="list",
    formula1='"New,Contacted,Interested,Converted,No Answer,Not Interested,Follow-up,Second Call Pending"',
    allow_blank=True, showErrorMessage=False
)
dv_status.sqref = f"F2:F10000"
ws.add_data_validation(dv_status)

# Note: Grade validation (G and P) is removed to allow multi-grade strings (e.g. '8, 9') without triggering 'Invalid' warnings.

# I – Second Call Done
dv_sc = DataValidation(
    type="list", formula1='"Yes,No"',
    allow_blank=True, showErrorMessage=False
)
dv_sc.sqref = "I2:I10000"
ws.add_data_validation(dv_sc)

# O – Paid
dv_paid = DataValidation(
    type="list", formula1='"Yes,No"',
    allow_blank=True, showErrorMessage=False
)
dv_paid.sqref = "O2:O10000"
ws.add_data_validation(dv_paid)

# D – Member dropdown (references Members sheet)
dv_member = DataValidation(
    type="list", formula1="Members!$A$2:$A$100",
    allow_blank=True, showErrorMessage=False
)
dv_member.sqref = "D2:D10000"
ws.add_data_validation(dv_member)

# ── Row 1 freeze + auto-filter ────────────────────────────────────────────────
ws.auto_filter.ref = "A1:P1"

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — Members
# ─────────────────────────────────────────────────────────────────────────────
ws_mem = wb.create_sheet("Members")
ws_mem.sheet_properties.tabColor = TEAL
ws_mem.sheet_view.showGridLines = True
ws_mem.freeze_panes = "A2"

mem_cols = {
    'A': ('Member Name', 22, NAVY),
    'B': ('Total Leads (auto)', 18, ROYAL),
    'C': ('Converted Leads (auto)', 20, GREEN),
    'D': ('Paid (auto)', 14, GREEN),
    'E': ('No Answer (auto)', 16, AMBER),
    'F': ('Conversion Rate (auto)', 20, TEAL),
    'G': ('Paid Rate (auto)', 18, TEAL),
    'H': ('Date Joined', 14, GRAY),
    'I': ('Notes / Role', 25, GRAY),
}

ws_mem.row_dimensions[1].height = 28
for col_letter, (header, width, bg) in mem_cols.items():
    idx = ord(col_letter) - ord('A') + 1
    ws_mem.column_dimensions[col_letter].width = width
    c = ws_mem.cell(1, idx)
    c.value = header
    c.fill  = pf(bg)
    c.font  = ft(10, True, WHITE)
    c.alignment = al("center", "center")
    c.border = med_bd()

# Member rows with formulas
ML = "'Master Leads'"
for mi, m in enumerate(MEMBERS, start=2):
    ws_mem.row_dimensions[mi].height = 22
    bg = LGREEN if mi % 2 == 0 else WHITE

    # A – Member name
    ws_mem.cell(mi, 1).value = m
    ws_mem.cell(mi, 1).font  = ft(10, True, NAVY)
    ws_mem.cell(mi, 1).fill  = pf(bg)
    ws_mem.cell(mi, 1).border = thin_bd()

    # B – Total Leads
    ws_mem.cell(mi, 2).value = f'=COUNTIF({ML}!D:D,A{mi})'
    ws_mem.cell(mi, 2).font  = ft(10, True, DGRAY)
    ws_mem.cell(mi, 2).fill  = pf(bg)
    ws_mem.cell(mi, 2).alignment = al("center")
    ws_mem.cell(mi, 2).border = thin_bd()

    # C – Converted
    ws_mem.cell(mi, 3).value = f'=COUNTIFS({ML}!D:D,A{mi},{ML}!F:F,"Converted")'
    ws_mem.cell(mi, 3).font  = ft(10, True, GREEN)
    ws_mem.cell(mi, 3).fill  = pf(bg)
    ws_mem.cell(mi, 3).alignment = al("center")
    ws_mem.cell(mi, 3).border = thin_bd()

    # D – Paid
    ws_mem.cell(mi, 4).value = f'=COUNTIFS({ML}!D:D,A{mi},{ML}!O:O,"Yes")'
    ws_mem.cell(mi, 4).font  = ft(10, True, GREEN)
    ws_mem.cell(mi, 4).fill  = pf(bg)
    ws_mem.cell(mi, 4).alignment = al("center")
    ws_mem.cell(mi, 4).border = thin_bd()

    # E – No Answer
    ws_mem.cell(mi, 5).value = f'=COUNTIFS({ML}!D:D,A{mi},{ML}!F:F,"No Answer")'
    ws_mem.cell(mi, 5).font  = ft(10, False, AMBER)
    ws_mem.cell(mi, 5).fill  = pf(bg)
    ws_mem.cell(mi, 5).alignment = al("center")
    ws_mem.cell(mi, 5).border = thin_bd()

    # F – Conversion Rate
    c_conv = ws_mem.cell(mi, 6)
    c_conv.value = f'=IFERROR(C{mi}/B{mi},"-")'
    c_conv.number_format = "0.0%"
    c_conv.font  = ft(10, False, DGRAY)
    c_conv.fill  = pf(bg)
    c_conv.alignment = al("center")
    c_conv.border = thin_bd()

    # G – Paid Rate
    c_paid = ws_mem.cell(mi, 7)
    c_paid.value = f'=IFERROR(D{mi}/B{mi},"-")'
    c_paid.number_format = "0.0%"
    c_paid.font  = ft(10, False, DGRAY)
    c_paid.fill  = pf(bg)
    c_paid.alignment = al("center")
    c_paid.border = thin_bd()

    # H – Date Joined (blank for user to fill)
    ws_mem.cell(mi, 8).fill = pf(bg); ws_mem.cell(mi, 8).border = thin_bd()
    ws_mem.cell(mi, 8).number_format = "YYYY-MM-DD"

    # I – Notes
    ws_mem.cell(mi, 9).fill = pf(bg); ws_mem.cell(mi, 9).border = thin_bd()

# Total row
tot_row = len(MEMBERS) + 2
ws_mem.row_dimensions[tot_row].height = 24
for c_idx in range(1, 10):
    ws_mem.cell(tot_row, c_idx).fill = pf(NAVY)
    ws_mem.cell(tot_row, c_idx).border = thin_bd()

ws_mem.cell(tot_row, 1).value = "TOTAL"
ws_mem.cell(tot_row, 1).font  = ft(10, True, WHITE)
ws_mem.cell(tot_row, 1).alignment = al("center")

ws_mem.cell(tot_row, 2).value = f'=COUNTA({ML}!B:B)-1'
ws_mem.cell(tot_row, 2).font  = ft(11, True, WHITE)
ws_mem.cell(tot_row, 2).alignment = al("center")

ws_mem.cell(tot_row, 3).value = f'=COUNTIF({ML}!F:F,"Converted")'
ws_mem.cell(tot_row, 3).font  = ft(11, True, WHITE)
ws_mem.cell(tot_row, 3).alignment = al("center")

ws_mem.cell(tot_row, 4).value = f'=COUNTIF({ML}!O:O,"Yes")'
ws_mem.cell(tot_row, 4).font  = ft(11, True, WHITE)
ws_mem.cell(tot_row, 4).alignment = al("center")

ws_mem.cell(tot_row, 5).value = f'=COUNTIF({ML}!F:F,"No Answer")'
ws_mem.cell(tot_row, 5).font  = ft(11, True, WHITE)
ws_mem.cell(tot_row, 5).alignment = al("center")

ws_mem.auto_filter.ref = "A1:I1"

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — Instructions (Google Sheets version)
# ─────────────────────────────────────────────────────────────────────────────
ws_inst = wb.create_sheet("Instructions")
ws_inst.sheet_properties.tabColor = GRAY
ws_inst.column_dimensions["A"].width = 90

instructions = [
    ("LEAD MANAGEMENT SYSTEM v9  —  Google Sheets Edition", NAVY, 18, True),
    ("", WHITE, 10, False),
    ("HOW TO USE THIS SYSTEM", ROYAL, 13, True),
    ("", WHITE, 10, False),
    ("1.  Upload this file to Google Sheets (File > Import, or drag and drop onto drive.google.com)", DGRAY, 10, False),
    ("2.  Go to Extensions > Apps Script and paste the new v9 Apps Script code", DGRAY, 10, False),
    ("3.  Save the script (Ctrl+S), then reload the sheet", DGRAY, 10, False),
    ("4.  A 'Lead Tools' menu will appear at the top", DGRAY, 10, False),
    ("", WHITE, 10, False),
    ("LEAD TOOLS MENU OPTIONS", ROYAL, 13, True),
    ("", WHITE, 10, False),
    ("  Add New Lead         — Enter phone + campaign + member. Auto: F-Code, Date, Duplicate check, Repeat tag", DGRAY, 10, False),
    ("  Look Up a Number     — See full history of any phone number across all campaigns", DGRAY, 10, False),
    ("  Add New Member       — Add a new team member to the Members sheet (auto-updates all dropdowns)", DGRAY, 10, False),
    ("  Remove Member        — Soft-remove: blanks the name but keeps all their lead history", DGRAY, 10, False),
    ("", WHITE, 10, False),
    ("COLUMNS IN MASTER LEADS", ROYAL, 13, True),
    ("", WHITE, 10, False),
    ("  A  Phone (Normalized)      — Auto-formula. Do NOT edit manually.", DGRAY, 10, False),
    ("  B  Raw Phone               — Enter exactly as dialled (+94 / 07xx / 9 digits)", DGRAY, 10, False),
    ("  C  F-Code                  — Auto-assigned. Same number always gets same F-Code.", DGRAY, 10, False),
    ("  D  Assigned Member         — Pick from dropdown (Members sheet auto-feeds this list)", DGRAY, 10, False),
    ("  E  Date Added              — Auto-filled when phone is entered", DGRAY, 10, False),
    ("  F  Status                  — Dropdown: New / Contacted / Interested / Converted / No Answer / etc.", DGRAY, 10, False),
    ("  G  Grade                   — Dropdown: 6, 7, 8, 9, 10, 11", DGRAY, 10, False),
    ("  H  Comments                — Free text notes", DGRAY, 10, False),
    ("  I  Second Call Done        — Dropdown: Yes / No", DGRAY, 10, False),
    ("  J  Second Call Notes       — Notes after second call", DGRAY, 10, False),
    ("  K  Duplicate Check         — Auto-formula. RED = same phone + same campaign already exists", DGRAY, 10, False),
    ("  L  Campaign / Boost Name   — e.g. B1, B2, Facebook July etc.", DGRAY, 10, False),
    ("  M  Repeat Student?         — Auto-formula. Detects if phone was seen in an earlier row.", DGRAY, 10, False),
    ("  N  Previous F-Code         — Auto-formula. Shows the original F-Code for repeat students.", DGRAY, 10, False),
    ("  O  Paid                    — Dropdown: Yes / No. Fill after fee is confirmed.", DGRAY, 10, False),
    ("  P  Grade (Final)           — Final enrolled grade (may differ from initial enquiry)", DGRAY, 10, False),
    ("", WHITE, 10, False),
    ("ADDING A NEW MEMBER", ROYAL, 13, True),
    ("", WHITE, 10, False),
    ("  Option 1 (Recommended): Lead Tools > Add New Member > type their name > Done!", DGRAY, 10, False),
    ("  Option 2 (Manual):      Go to Members sheet > type name in column A below last entry", DGRAY, 10, False),
    ("  Both methods: The member dropdown in Master Leads col D will auto-show them immediately.", DGRAY, 10, False),
    ("", WHITE, 10, False),
    ("COLOUR CODING IN MASTER LEADS", ROYAL, 13, True),
    ("", WHITE, 10, False),
    ("  RED row    — Duplicate (same phone + same campaign already logged)", DGRAY, 10, False),
    ("  ORANGE     — Phone entered but F-Code or Member not yet filled", DGRAY, 10, False),
    ("  GREEN      — Status = Converted", DGRAY, 10, False),
    ("  BLUE       — Status = New", DGRAY, 10, False),
    ("  GREY       — Status = Contacted", DGRAY, 10, False),
    ("  LIGHT BLUE — Repeat Student row", DGRAY, 10, False),
]

for ri, (text, bg, size, bold) in enumerate(instructions, start=1):
    ws_inst.row_dimensions[ri].height = 20 if text else 8
    c = ws_inst.cell(ri, 1)
    c.value = text
    c.fill  = pf(bg)
    c.font  = ft(size, bold, WHITE if bg != WHITE else DGRAY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(DST)
print(f"Clean Google Sheets file saved: {DST}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
