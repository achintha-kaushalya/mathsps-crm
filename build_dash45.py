#!/usr/bin/env python3
"""
Lead CRM - Dashboard 4: Member Performance  +  Dashboard 5: Member Paid
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Lead_CRM_Dashboards.xlsx'
ML  = "'Master Leads'"

NAVY   = "0D2744"; BLUE   = "1565C0"; ROYAL  = "1976D2"; SKY    = "0EA5E9"
WHITE  = "FFFFFF"; OFF_W  = "F0F4F8"; LGRAY  = "E2E8F0"; GRAY   = "64748B"
DGRAY  = "1E293B"; GREEN  = "166534"; LGREEN = "DCFCE7"; TEAL   = "134E4A"
LTEAL  = "CCFBF1"; AMBER  = "C2410C"; LAMBER = "FFF7ED"; RED    = "991B1B"
PURPLE = "581C87"; LPURPLE= "F3E8FF"; INDIGO = "312E81"; LINDIGO= "E0E7FF"
CARD_BG= "FFFFFF"

GRADE_COLS = ["6","7","8","9","10","11"]
MEMBER_ACCENT_COLORS = ["1D4ED8","047857","7C3AED","B45309","0E7490","BE185D","166534","C2410C"]

def pf(c):   return PatternFill("solid", fgColor=c)
def ft(size=10, bold=False, color="000000", name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_s(color="D1D5DB"):
    return Side(border_style="thin", color=color)
def bd_all(color="D1D5DB"):
    s = thin_s(color)
    return Border(left=s, right=s, top=s, bottom=s)
def bd_bottom(color="D1D5DB"):
    s = thin_s(color)
    n = Side(border_style=None)
    return Border(left=n, right=n, top=n, bottom=s)

def merge_set(ws, r1, c1, r2, c2, val=None, fill_color=None,
              font=None, align=None, numfmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if val is not None: cell.value = val
    if font:            cell.font  = font
    if align:           cell.alignment = align
    if numfmt:          cell.number_format = numfmt
    if fill_color:
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).fill = pf(fill_color)
    return cell

def section_hdr(ws, row, c1, c2, label, bg=LGRAY):
    for c in range(1, c2+2): ws.cell(row, c).fill = pf(bg)
    merge_set(ws, row, c1, row, c2, val=label, fill_color=bg,
              font=ft(10, True, DGRAY), align=al("left", "center"))

def setup_sheet(ws, tab_color, title_text, subtitle_text, title_bg, sub_bg, sub_text_color):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.column_dimensions["A"].width = 1.5
    for c in range(2, 40): ws.column_dimensions[get_column_letter(c)].width = 3.5
    for c in range(40, 80): ws.column_dimensions[get_column_letter(c)].width = 13
    for r, h in {1:7, 2:52, 3:20, 4:14}.items():
        ws.row_dimensions[r].height = h
    for r in range(5, 90): ws.row_dimensions[r].height = 14
    for r in range(1, 85):
        for c in range(1, 38): ws.cell(r, c).fill = pf(OFF_W)
    for c in range(1, 38): ws.cell(2, c).fill = pf(title_bg)
    merge_set(ws, 2, 2, 2, 36, val=title_text, fill_color=title_bg,
              font=ft(20, True, WHITE), align=al("left", "center"))
    for c in range(1, 38): ws.cell(3, c).fill = pf(sub_bg)
    merge_set(ws, 3, 2, 3, 36, val=subtitle_text, fill_color=sub_bg,
              font=ft(9, False, sub_text_color), align=al("left", "center"))

# ═══════════════════════════════════════════════════════════════════════════════
# LOAD & GET MEMBERS LIST
# ═══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(DST)
mem_ws = wb["Members"]
members = []
for r in range(2, 50):
    v = mem_ws.cell(r, 1).value
    if v: members.append(str(v).strip())
    else: break

print(f"Members found: {members}")
N = len(members)

# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD 4 – MEMBER PERFORMANCE
# ═══════════════════════════════════════════════════════════════════════════════
H4 = "_DB4_Data"
if H4 in wb.sheetnames: del wb[H4]
hws4 = wb.create_sheet(H4); hws4.sheet_state = "hidden"

D4 = "Member Performance"
if D4 in wb.sheetnames: del wb[D4]
ws4 = wb.create_sheet(D4, 3)

setup_sheet(ws4, INDIGO,
    "  MEMBER PERFORMANCE DASHBOARD   |   Staff Leads, Conversions & Rankings",
    "  See which team member handles the most leads and converts best  \u00b7  Auto-updates",
    INDIGO, "1E3A8A", "BFDBFE")

# ── Member Performance Table ──────────────────────────────────────────────────
section_hdr(ws4, 5, 2, 36, "  MEMBER-WISE LEAD & CONVERSION PERFORMANCE")
ws4.row_dimensions[5].height = 26

# Table header (row 6)
ws4.row_dimensions[6].height = 22
hdrs = [("Member", 5, NAVY), ("Total Leads", 4, BLUE), ("Paid", 4, GREEN),
        ("Converted", 4, TEAL), ("No Answer", 4, AMBER), ("Conv. %", 4, PURPLE),
        ("Paid %", 4, INDIGO), ("Rank", 3, RED)]
col = 2
for hdr_text, span, hcol in hdrs:
    merge_set(ws4, 6, col, 6, col+span-1, val=hdr_text, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

# Member rows
for mi, m in enumerate(members):
    r = 7 + mi
    ws4.row_dimensions[r].height = 20
    bg = LINDIGO if mi % 2 == 0 else WHITE
    acc = MEMBER_ACCENT_COLORS[mi % len(MEMBER_ACCENT_COLORS)]

    col = 2
    # Member name
    merge_set(ws4, r, col, r, col+4, val=m, fill_color=acc,
              font=ft(10, True, WHITE), align=al("left", "center")); col += 5

    # Total Leads
    leads_f = f'=COUNTIF({ML}!D:D,"{m}")'
    merge_set(ws4, r, col, r, col+3, val=leads_f, fill_color=bg,
              font=ft(10, True, DGRAY), align=al("center", "center")); col += 4

    # Paid
    paid_f = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!O:O,"Yes")'
    merge_set(ws4, r, col, r, col+3, val=paid_f, fill_color=bg,
              font=ft(10, True, GREEN), align=al("center", "center")); col += 4

    # Converted
    conv_f = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!F:F,"Converted")'
    merge_set(ws4, r, col, r, col+3, val=conv_f, fill_color=bg,
              font=ft(10, True, TEAL), align=al("center", "center")); col += 4

    # No Answer
    na_f = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!F:F,"No Answer")'
    merge_set(ws4, r, col, r, col+3, val=na_f, fill_color=bg,
              font=ft(10, False, AMBER), align=al("center", "center")); col += 4

    # Conv %
    convpct_f = f'=IFERROR(COUNTIFS({ML}!D:D,"{m}",{ML}!F:F,"Converted")/COUNTIF({ML}!D:D,"{m}"),0)'
    c1 = merge_set(ws4, r, col, r, col+3, val=convpct_f, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c1.number_format = "0.0%"; col += 4

    # Paid %
    paidpct_f = f'=IFERROR(COUNTIFS({ML}!D:D,"{m}",{ML}!O:O,"Yes")/COUNTIF({ML}!D:D,"{m}"),0)'
    c2 = merge_set(ws4, r, col, r, col+3, val=paidpct_f, fill_color=bg,
              font=ft(10, False, DGRAY), align=al("center", "center"))
    c2.number_format = "0.0%"; col += 4

    # Rank (RANK by leads)
    rank_f = f'=IFERROR(RANK(COUNTIF({ML}!D:D,"{m}"),COUNTIF({ML}!D:D,"{{{",".join(members)}"}})),"-")'
    # Simplified rank using RANK on the leads value
    leads_val_ref = f'COUNTIF({ML}!D:D,"{m}")'
    all_leads_arr = ",".join([f'COUNTIF({ML}!D:D,"{x}")' for x in members])
    rank_f2 = f'=RANK(COUNTIF({ML}!D:D,"{m}"),{{{all_leads_arr}}})'
    merge_set(ws4, r, col, r, col+2, val=rank_f2, fill_color=bg,
              font=ft(10, True, PURPLE), align=al("center", "center"))

# Total row
r_tot = 7 + N
ws4.row_dimensions[r_tot].height = 22
for c in range(2, 37): ws4.cell(r_tot, c).fill = pf(NAVY)
merge_set(ws4, r_tot, 2, r_tot, 6, val="TOTAL", fill_color=NAVY,
          font=ft(10, True, WHITE), align=al("center", "center"))
merge_set(ws4, r_tot, 7, r_tot, 10,
          val=f"=COUNTA({ML}!B:B)-1", fill_color=NAVY,
          font=ft(11, True, WHITE), align=al("center", "center"))
merge_set(ws4, r_tot, 11, r_tot, 14,
          val=f'=COUNTIF({ML}!O:O,"Yes")', fill_color=NAVY,
          font=ft(11, True, WHITE), align=al("center", "center"))

# ── Grade Breakdown per Member ────────────────────────────────────────────────
gbr = r_tot + 2
ws4.row_dimensions[gbr].height = 26
section_hdr(ws4, gbr, 2, 36, "  GRADE BREAKDOWN BY MEMBER   —   Which member handles which grade")

# Header row
ws4.row_dimensions[gbr+1].height = 22
grade_hdr_cols = [("Member", 5, NAVY)] + [(f"Grade {g}", 3, MEMBER_ACCENT_COLORS[gi]) for gi, g in enumerate(GRADE_COLS)] + [("Total", 3, DGRAY)]
col = 2
for hdr_text, span, hcol in grade_hdr_cols:
    merge_set(ws4, gbr+1, col, gbr+1, col+span-1, val=hdr_text, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

for mi, m in enumerate(members):
    r = gbr + 2 + mi
    ws4.row_dimensions[r].height = 20
    bg = LTEAL if mi % 2 == 0 else WHITE
    acc = MEMBER_ACCENT_COLORS[mi % len(MEMBER_ACCENT_COLORS)]
    col = 2
    merge_set(ws4, r, col, r, col+4, val=m, fill_color=acc,
              font=ft(10, True, WHITE), align=al("left", "center")); col += 5
    for g in GRADE_COLS:
        f_grade = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!G:G,{g})'
        merge_set(ws4, r, col, r, col+2, val=f_grade, fill_color=bg,
                  font=ft(10, False, DGRAY), align=al("center", "center")); col += 3
    total_f = f'=COUNTIF({ML}!D:D,"{m}")'
    merge_set(ws4, r, col, r, col+2, val=total_f, fill_color=bg,
              font=ft(10, True, NAVY), align=al("center", "center"))

# ── Helper data for charts ────────────────────────────────────────────────────
hws4.cell(1,1).value = "Member"; hws4.cell(1,2).value = "Total Leads"
hws4.cell(1,3).value = "Paid";   hws4.cell(1,4).value = "Converted"
for mi, m in enumerate(members):
    hws4.cell(2+mi, 1).value = m
    hws4.cell(2+mi, 2).value = f'=COUNTIF({ML}!D:D,"{m}")'
    hws4.cell(2+mi, 3).value = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!O:O,"Yes")'
    hws4.cell(2+mi, 4).value = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!F:F,"Converted")'

# ── Charts ────────────────────────────────────────────────────────────────────
chart_row = gbr + 2 + N + 2
ws4.row_dimensions[chart_row].height = 26
section_hdr(ws4, chart_row, 2, 36, "  MEMBER CHARTS   —   Visual performance comparison")

cat_ref  = Reference(hws4, min_col=1, max_col=1, min_row=2, max_row=1+N)
leads_ref = Reference(hws4, min_col=2, max_col=2, min_row=1, max_row=1+N)
paid_ref  = Reference(hws4, min_col=3, max_col=3, min_row=1, max_row=1+N)
conv_ref  = Reference(hws4, min_col=4, max_col=4, min_row=1, max_row=1+N)

# Bar: Leads handled
bar_leads = BarChart()
bar_leads.type = "bar"; bar_leads.title = "Leads Handled per Member"
bar_leads.style = 10; bar_leads.legend = None
bar_leads.y_axis.title = "Member"; bar_leads.x_axis.title = "Leads"
bar_leads.add_data(leads_ref, titles_from_data=True)
bar_leads.set_categories(cat_ref)
bar_leads.series[0].graphicalProperties.solidFill = INDIGO
bar_leads.width = 14; bar_leads.height = 10
ws4.add_chart(bar_leads, "B" + str(chart_row + 1))

# Bar: Paid handled
bar_paid = BarChart()
bar_paid.type = "bar"; bar_paid.title = "Paid Students per Member"
bar_paid.style = 10; bar_paid.legend = None
bar_paid.y_axis.title = "Member"; bar_paid.x_axis.title = "Paid"
bar_paid.add_data(paid_ref, titles_from_data=True)
bar_paid.set_categories(cat_ref)
bar_paid.series[0].graphicalProperties.solidFill = GREEN
bar_paid.width = 14; bar_paid.height = 10
ws4.add_chart(bar_paid, "L" + str(chart_row + 1))

# Clustered bar: Leads + Converted
bar_comp = BarChart()
bar_comp.type = "bar"; bar_comp.title = "Leads vs Converted per Member"
bar_comp.style = 10; bar_comp.grouping = "clustered"
bar_comp.y_axis.title = "Member"; bar_comp.x_axis.title = "Count"
bar_comp.add_data(leads_ref, titles_from_data=True)
bar_comp.add_data(conv_ref,  titles_from_data=True)
bar_comp.set_categories(cat_ref)
bar_comp.series[0].graphicalProperties.solidFill = INDIGO
bar_comp.series[1].graphicalProperties.solidFill = TEAL
bar_comp.width = 14; bar_comp.height = 10
ws4.add_chart(bar_comp, "T" + str(chart_row + 1))

# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD 5 – MEMBER PAID
# ═══════════════════════════════════════════════════════════════════════════════
H5 = "_DB5_Data"
if H5 in wb.sheetnames: del wb[H5]
hws5 = wb.create_sheet(H5); hws5.sheet_state = "hidden"

D5 = "Member Paid"
if D5 in wb.sheetnames: del wb[D5]
ws5 = wb.create_sheet(D5, 4)

setup_sheet(ws5, GREEN,
    "  MEMBER PAID DASHBOARD   |   Paid Students by Member & Grade",
    "  Drill down into which member generated paid students in each grade  \u00b7  Auto-updates",
    GREEN, TEAL, "A7F3D0")

# ── Paid per Member × Grade Table ────────────────────────────────────────────
section_hdr(ws5, 5, 2, 36, "  PAID BY MEMBER \u00d7 GRADE   —   Detailed grade-wise paid breakdown per member")
ws5.row_dimensions[5].height = 26

# Header
ws5.row_dimensions[6].height = 22
hdrs5 = [("Member", 5, NAVY)] + [(f"Grade {g}", 3, MEMBER_ACCENT_COLORS[gi]) for gi, g in enumerate(GRADE_COLS)] + [("Total Paid", 4, GREEN)]
col = 2
for hdr_text, span, hcol in hdrs5:
    merge_set(ws5, 6, col, 6, col+span-1, val=hdr_text, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

for mi, m in enumerate(members):
    r = 7 + mi
    ws5.row_dimensions[r].height = 20
    bg = LGREEN if mi % 2 == 0 else WHITE
    acc = MEMBER_ACCENT_COLORS[mi % len(MEMBER_ACCENT_COLORS)]
    col = 2
    merge_set(ws5, r, col, r, col+4, val=m, fill_color=acc,
              font=ft(10, True, WHITE), align=al("left", "center")); col += 5
    for g in GRADE_COLS:
        f_gp = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!G:G,{g},{ML}!O:O,"Yes")'
        merge_set(ws5, r, col, r, col+2, val=f_gp, fill_color=bg,
                  font=ft(10, False, DGRAY), align=al("center", "center")); col += 3
    total_paid_f = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!O:O,"Yes")'
    merge_set(ws5, r, col, r, col+3, val=total_paid_f, fill_color=bg,
              font=ft(10, True, GREEN), align=al("center", "center"))

# Total row
r_tot5 = 7 + N
ws5.row_dimensions[r_tot5].height = 22
for c in range(2, 36): ws5.cell(r_tot5, c).fill = pf(NAVY)
merge_set(ws5, r_tot5, 2, r_tot5, 6, val="TOTAL", fill_color=NAVY,
          font=ft(10, True, WHITE), align=al("center", "center"))
col = 7
for g in GRADE_COLS:
    merge_set(ws5, r_tot5, col, r_tot5, col+2,
              val=f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")',
              fill_color=NAVY, font=ft(10, True, WHITE), align=al("center", "center"))
    col += 3
merge_set(ws5, r_tot5, col, r_tot5, col+3,
          val=f'=COUNTIF({ML}!O:O,"Yes")',
          fill_color=NAVY, font=ft(10, True, WHITE), align=al("center", "center"))

# ── Helper data for Charts ────────────────────────────────────────────────────
hws5.cell(1,1).value = "Member"
for gi, g in enumerate(GRADE_COLS):
    hws5.cell(1, 2+gi).value = f"Grade {g}"
    hws5.cell(1, 2+gi).font = Font(bold=True)
hws5.cell(1, 8).value = "Total Paid"

for mi, m in enumerate(members):
    hws5.cell(2+mi, 1).value = m
    for gi, g in enumerate(GRADE_COLS):
        hws5.cell(2+mi, 2+gi).value = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!G:G,{g},{ML}!O:O,"Yes")'
    hws5.cell(2+mi, 8).value = f'=COUNTIFS({ML}!D:D,"{m}",{ML}!O:O,"Yes")'

# Grade totals row for grade chart
hws5.cell(2+N, 1).value = "Total"
for gi, g in enumerate(GRADE_COLS):
    hws5.cell(2+N, 2+gi).value = f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")'
hws5.cell(2+N, 8).value = f'=COUNTIF({ML}!O:O,"Yes")'

# ── Charts ────────────────────────────────────────────────────────────────────
chart_r5 = r_tot5 + 2
ws5.row_dimensions[chart_r5].height = 26
section_hdr(ws5, chart_r5, 2, 36, "  PAID CHARTS   —   Paid breakdown by member and grade")

# Bar: Paid by Member
bar5a = BarChart()
bar5a.type = "bar"; bar5a.title = "Total Paid by Member"
bar5a.style = 10; bar5a.legend = None
bar5a.y_axis.title = "Member"; bar5a.x_axis.title = "Paid Students"
cat5 = Reference(hws5, min_col=1, max_col=1, min_row=2, max_row=1+N)
val5a = Reference(hws5, min_col=8, max_col=8, min_row=1, max_row=1+N)
bar5a.add_data(val5a, titles_from_data=True)
bar5a.set_categories(cat5)
bar5a.series[0].graphicalProperties.solidFill = GREEN
bar5a.width = 14; bar5a.height = 11
ws5.add_chart(bar5a, "B" + str(chart_r5+1))

# Bar: Paid by Grade (totals)
bar5b = BarChart()
bar5b.type = "col"; bar5b.title = "Paid Students by Grade"
bar5b.style = 10; bar5b.legend = None
bar5b.x_axis.title = "Grade"; bar5b.y_axis.title = "Paid"
grade_cats_5 = Reference(hws5, min_col=2, max_col=7, min_row=1, max_row=1)
grade_vals_5 = Reference(hws5, min_col=2, max_col=7, min_row=2+N, max_row=2+N)
bar5b.add_data(grade_vals_5, titles_from_data=False)
bar5b.set_categories(grade_cats_5)
bar5b.series[0].graphicalProperties.solidFill = TEAL
bar5b.width = 14; bar5b.height = 11
ws5.add_chart(bar5b, "L" + str(chart_r5+1))

# Stacked bar: Paid by Member × Grade (multi-series)
stack5 = BarChart()
stack5.type = "bar"; stack5.title = "Paid by Member \u00d7 Grade"
stack5.style = 10; stack5.grouping = "stacked"
stack5.y_axis.title = "Member"; stack5.x_axis.title = "Paid Students"
grade_colors_5 = ["1D4ED8","0369A1","047857","7C3AED","B45309","B91C1C"]
cat5s = Reference(hws5, min_col=1, max_col=1, min_row=2, max_row=1+N)
for gi in range(6):
    v5 = Reference(hws5, min_col=2+gi, max_col=2+gi, min_row=1, max_row=1+N)
    stack5.add_data(v5, titles_from_data=True)
    stack5.series[gi].graphicalProperties.solidFill = grade_colors_5[gi]
stack5.set_categories(cat5s)
stack5.width = 14; stack5.height = 11
ws5.add_chart(stack5, "T" + str(chart_r5+1))

# ═══════════════════════════════════════════════════════════════════════════════
wb.save(DST)
print("Dashboard 4 (Member Performance) & Dashboard 5 (Member Paid) saved.")
print("Sheets:", wb.sheetnames)
