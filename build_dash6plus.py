#!/usr/bin/env python3
"""
Lead CRM - Dashboards 6, 7, 8 + Repeat + Follow-up + Revenue Shell
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Lead_CRM_Dashboards.xlsx'
ML  = "'Master Leads'"

NAVY   = "0D2744"; BLUE   = "1565C0"; ROYAL  = "1976D2"; SKY    = "0EA5E9"
WHITE  = "FFFFFF"; OFF_W  = "F0F4F8"; LGRAY  = "E2E8F0"; GRAY   = "64748B"
DGRAY  = "1E293B"; GREEN  = "166534"; LGREEN = "DCFCE7"; TEAL   = "134E4A"
LTEAL  = "CCFBF1"; AMBER  = "C2410C"; LAMBER = "FFF7ED"; RED    = "991B1B"
LRED   = "FEE2E2"; PURPLE = "581C87"; LPURPLE= "F3E8FF"; INDIGO = "312E81"
LINDIGO= "E0E7FF"; CARD_BG= "FFFFFF"; ORANGE = "EA580C"; LORANGE= "FFF7ED"
ROSE   = "9F1239"; LROSE  = "FFF1F2"

MEMBER_ACCENT_COLORS = ["1D4ED8","047857","7C3AED","B45309","0E7490","BE185D","166534","C2410C"]

def pf(c): return PatternFill("solid", fgColor=c)
def ft(size=10, bold=False, color="000000", name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_bd(color="D1D5DB"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def merge_set(ws, r1, c1, r2, c2, val=None, fill_color=None,
              font=None, align=None, numfmt=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1)
    if val is not None: cell.value = val
    if font:            cell.font  = font
    if align:           cell.alignment = align
    if numfmt:          cell.number_format = numfmt
    if fill_color:
        for r in range(r1, r2+1):
            for c in range(c1, c2+1):
                ws.cell(r, c).fill = pf(fill_color)
    return cell

def section_hdr(ws, row, c1, c2, label, bg=LGRAY):
    for c in range(1, c2+2): ws.cell(row, c).fill = pf(bg)
    merge_set(ws, row, c1, row, c2, val=label, fill_color=bg,
              font=ft(10, True, DGRAY), align=al("left", "center"))

def kpi_card(ws, row, col, title, formula, sublabel, accent, W=4, numfmt=None):
    for r in range(row, row+3):
        for c in range(col, col+W):
            ws.cell(r, c).fill = pf(accent if r == row else CARD_BG)
    acc_s  = Side(border_style="medium", color=accent)
    gray_s = Side(border_style="thin",   color="D1D5DB")
    none_s = Side(border_style=None)
    for r in range(row, row+3):
        for c in range(col, col+W):
            ws.cell(r, c).border = Border(
                left=acc_s if c == col else none_s,
                right=gray_s if c == col+W-1 else none_s,
                top=gray_s if r == row else none_s,
                bottom=gray_s if r == row+2 else none_s
            )
    merge_set(ws, row, col, row, col+W-1, val=title.upper(), fill_color=accent,
              font=ft(8, True, WHITE), align=al("left", "center"))
    vc = merge_set(ws, row+1, col, row+1, col+W-1, val=formula, fill_color=CARD_BG,
              font=ft(20, True, accent), align=al("left", "center"))
    if numfmt: vc.number_format = numfmt
    merge_set(ws, row+2, col, row+2, col+W-1, val=sublabel, fill_color=CARD_BG,
              font=ft(8, False, GRAY), align=al("left", "top"))

def setup_sheet(ws, tab_color, title_text, subtitle_text, tbg, sbg, stc, ncols=30):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False; ws.sheet_view.zoomScale = 85
    ws.column_dimensions["A"].width = 1.5
    for c in range(2, ncols+5): ws.column_dimensions[get_column_letter(c)].width = 3.7
    for c in range(ncols+5, 80): ws.column_dimensions[get_column_letter(c)].width = 13
    for r, h in {1:7, 2:52, 3:20, 4:14}.items(): ws.row_dimensions[r].height = h
    for r in range(5, 90): ws.row_dimensions[r].height = 14
    for r in range(1, 85):
        for c in range(1, ncols+3): ws.cell(r, c).fill = pf(OFF_W)
    for c in range(1, ncols+3): ws.cell(2, c).fill = pf(tbg)
    merge_set(ws, 2, 2, 2, ncols+1, val=title_text, fill_color=tbg,
              font=ft(20, True, WHITE), align=al("left", "center"))
    for c in range(1, ncols+3): ws.cell(3, c).fill = pf(sbg)
    merge_set(ws, 3, 2, 3, ncols+1, val=subtitle_text, fill_color=sbg,
              font=ft(9, False, stc), align=al("left", "center"))

# ═══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(DST)
mem_ws = wb["Members"]
members = []
for r in range(2, 50):
    v = mem_ws.cell(r, 1).value
    if v: members.append(str(v).strip())
    else: break
N = len(members)
GRADES = ["6","7","8","9","10","11"]

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  DASHBOARD 6 – CAMPAIGN DASHBOARD                                       ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
H6 = "_DB6_Data"
if H6 in wb.sheetnames: del wb[H6]
hws6 = wb.create_sheet(H6); hws6.sheet_state = "hidden"

D6 = "Campaign Dashboard"
if D6 in wb.sheetnames: del wb[D6]
ws6 = wb.create_sheet(D6, 5)

setup_sheet(ws6, AMBER,
    "  CAMPAIGN DASHBOARD   |   Lead & Paid Performance by Campaign / Boost",
    "  Track which campaigns bring the most leads and conversions  \u00b7  Add campaign names in _DB6_Data rows 2-21",
    AMBER, ORANGE, "FED7AA")

# Campaigns list (user can extend _DB6_Data)
campaigns = ["B1","B2","B3","B4","B5","B6","B7","B8","B9","B10",
             "B11","B12","B13","B14","B15","B16","B17","B18","B19","B20"]

# Helper data
hws6.cell(1,1).value="Campaign"; hws6.cell(1,2).value="Leads"
hws6.cell(1,3).value="Paid"; hws6.cell(1,4).value="Converted"; hws6.cell(1,5).value="Conv%"
for i, camp in enumerate(campaigns):
    r6 = 2 + i
    hws6.cell(r6,1).value = camp
    hws6.cell(r6,2).value = f'=COUNTIF({ML}!L:L,"{camp}")'
    hws6.cell(r6,3).value = f'=COUNTIFS({ML}!L:L,"{camp}",{ML}!O:O,"Yes")'
    hws6.cell(r6,4).value = f'=COUNTIFS({ML}!L:L,"{camp}",{ML}!F:F,"Converted")'
    hws6.cell(r6,5).value = f'=IFERROR(COUNTIFS({ML}!L:L,"{camp}",{ML}!F:F,"Converted")/COUNTIF({ML}!L:L,"{camp}"),0)'
    hws6.cell(r6,5).number_format = "0.0%"

# KPI Cards
section_hdr(ws6, 5, 2, 30, "  CAMPAIGN OVERVIEW KPIs")
ws6.row_dimensions[5].height = 26
for r in range(5,16): ws6.row_dimensions[r].height = 14
ws6.row_dimensions[6].height = 14; ws6.row_dimensions[7].height = 38; ws6.row_dimensions[8].height = 15

kpis6 = [
    (2,  "TOTAL CAMPAIGNS",
     f'=SUMPRODUCT((COUNTIF({ML}!L:L,{{"B1","B2","B3","B4","B5","B6","B7","B8","B9","B10","B11","B12","B13","B14","B15","B16","B17","B18","B19","B20"}})>0)*1)',
     "Active campaigns (B1-B20)", AMBER, None),
    (7,  "TOTAL LEADS",
     f"=COUNTA({ML}!B:B)-1",
     "All leads across all campaigns", NAVY, None),
    (12, "TOTAL PAID",
     f'=COUNTIF({ML}!O:O,"Yes")',
     "Paid across all campaigns", GREEN, None),
    (17, "TOTAL CONVERTED",
     f'=COUNTIF({ML}!F:F,"Converted")',
     "Converted across all campaigns", TEAL, None),
    (22, "LEADS WITH NO CAMPAIGN",
     f'=COUNTIF({ML}!L:L,"")+"COUNTBLANK"',
     "Leads missing campaign name", ROSE, None),
]

# Fix the last KPI
kpis6[-1] = (22, "NO CAMPAIGN ASSIGNED",
    f'=COUNTBLANK({ML}!L2:L10000)',
    "Leads missing campaign name", ROSE, None)

for col, title, formula, sublabel, accent, numfmt in kpis6:
    kpi_card(ws6, 6, col, title, formula, sublabel, accent, numfmt=numfmt)

# Campaign table
section_hdr(ws6, 11, 2, 30, "  CAMPAIGN-WISE BREAKDOWN   —   Leads, Paid, Converted, Conversion %")
ws6.row_dimensions[11].height = 26
ws6.row_dimensions[12].height = 22

# Table headers
hdrs6 = [("Campaign", 4, AMBER), ("Leads", 4, NAVY), ("Paid", 4, GREEN),
          ("Converted", 4, TEAL), ("Conv. %", 4, PURPLE), ("Rank", 3, RED)]
col = 2
for ht, span, hcol in hdrs6:
    merge_set(ws6, 12, col, 12, col+span-1, val=ht, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

for i, camp in enumerate(campaigns):
    r6 = 13 + i
    ws6.row_dimensions[r6].height = 18
    bg = LAMBER if i % 2 == 0 else WHITE
    col = 2
    merge_set(ws6, r6, col, r6, col+3, val=camp, fill_color=AMBER,
              font=ft(10, True, WHITE), align=al("center", "center")); col += 4
    merge_set(ws6, r6, col, r6, col+3,
              val=f'=COUNTIF({ML}!L:L,"{camp}")', fill_color=bg,
              font=ft(10, True, DGRAY), align=al("center", "center")); col += 4
    merge_set(ws6, r6, col, r6, col+3,
              val=f'=COUNTIFS({ML}!L:L,"{camp}",{ML}!O:O,"Yes")', fill_color=bg,
              font=ft(10, True, GREEN), align=al("center", "center")); col += 4
    merge_set(ws6, r6, col, r6, col+3,
              val=f'=COUNTIFS({ML}!L:L,"{camp}",{ML}!F:F,"Converted")', fill_color=bg,
              font=ft(10, True, TEAL), align=al("center", "center")); col += 4
    pct_c = merge_set(ws6, r6, col, r6, col+3,
              val=f'=IFERROR(COUNTIFS({ML}!L:L,"{camp}",{ML}!F:F,"Converted")/COUNTIF({ML}!L:L,"{camp}"),0)',
              fill_color=bg, font=ft(10, False, DGRAY), align=al("center", "center"))
    pct_c.number_format = "0.0%"; col += 4
    # Rank by leads
    all_camp_leads = ",".join([f'COUNTIF({ML}!L:L,"{c}")' for c in campaigns])
    rnk = f'=IFERROR(RANK(COUNTIF({ML}!L:L,"{camp}"),{{{all_camp_leads}}}),"-")'
    merge_set(ws6, r6, col, r6, col+2, val=rnk, fill_color=bg,
              font=ft(10, True, RED), align=al("center", "center"))

# Charts
chart_r6 = 13 + len(campaigns) + 2
ws6.row_dimensions[chart_r6].height = 26
section_hdr(ws6, chart_r6, 2, 30, "  CAMPAIGN CHARTS")

cat6  = Reference(hws6, min_col=1, max_col=1, min_row=2, max_row=11)
ld6   = Reference(hws6, min_col=2, max_col=2, min_row=1, max_row=11)
pd6   = Reference(hws6, min_col=3, max_col=3, min_row=1, max_row=11)

bar6a = BarChart(); bar6a.type = "bar"; bar6a.title = "Campaign Leads & Paid"
bar6a.style = 10; bar6a.grouping = "clustered"
bar6a.y_axis.title = "Campaign"; bar6a.x_axis.title = "Count"
bar6a.add_data(ld6, titles_from_data=True)
bar6a.add_data(pd6, titles_from_data=True)
bar6a.set_categories(cat6)
bar6a.series[0].graphicalProperties.solidFill = AMBER
bar6a.series[1].graphicalProperties.solidFill = GREEN
bar6a.width = 22; bar6a.height = 11
ws6.add_chart(bar6a, "B" + str(chart_r6+1))

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  DASHBOARD 7 – DAILY DASHBOARD                                          ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
H7 = "_DB7_Data"
if H7 in wb.sheetnames: del wb[H7]
hws7 = wb.create_sheet(H7); hws7.sheet_state = "hidden"

D7 = "Daily Dashboard"
if D7 in wb.sheetnames: del wb[D7]
ws7 = wb.create_sheet(D7, 6)

setup_sheet(ws7, SKY,
    "  DAILY DASHBOARD   |   Today, Yesterday, Last 7 Days, This Week, This Month",
    "  Quick snapshot of lead activity over different time periods  \u00b7  Auto-refreshes daily",
    SKY, "0284C7", "BAE6FD")

# Time-period KPI Cards
for r in range(5, 14): ws7.row_dimensions[r].height = 14
ws7.row_dimensions[6].height = 14; ws7.row_dimensions[7].height = 38; ws7.row_dimensions[8].height = 15
ws7.row_dimensions[10].height = 14; ws7.row_dimensions[11].height = 38; ws7.row_dimensions[12].height = 15

section_hdr(ws7, 5, 2, 30, "  TIME-PERIOD LEAD COUNTS   —   Auto-refreshes on every open")
ws7.row_dimensions[5].height = 26

kpis7_r1 = [
    (2,  "TODAY",
     f"=COUNTIF({ML}!E:E,TODAY())",
     "Leads added today", NAVY, None),
    (7,  "YESTERDAY",
     f"=COUNTIF({ML}!E:E,TODAY()-1)",
     "Leads added yesterday", BLUE, None),
    (12, "LAST 7 DAYS",
     f'=COUNTIFS({ML}!E:E,">="&TODAY()-6,{ML}!E:E,"<="&TODAY())',
     "Leads in the last 7 days", ROYAL, None),
    (17, "THIS WEEK",
     f'=COUNTIFS({ML}!E:E,">="&TODAY()-WEEKDAY(TODAY(),2)+1,{ML}!E:E,"<="&TODAY())',
     "Mon to today this week", SKY, None),
    (22, "THIS MONTH",
     f'=COUNTIFS({ML}!E:E,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),{ML}!E:E,"<="&TODAY())',
     "First of month to today", TEAL, None),
]

kpis7_r2 = [
    (2,  "LAST MONTH",
     f'=COUNTIFS({ML}!E:E,">="&DATE(YEAR(TODAY()),MONTH(TODAY())-1,1),{ML}!E:E,"<="&EOMONTH(TODAY(),-1))',
     "All of last month", PURPLE, None),
    (7,  "THIS WEEK PAID",
     f'=COUNTIFS({ML}!E:E,">="&TODAY()-WEEKDAY(TODAY(),2)+1,{ML}!E:E,"<="&TODAY(),{ML}!O:O,"Yes")',
     "Paid this week", GREEN, None),
    (12, "THIS MONTH PAID",
     f'=COUNTIFS({ML}!E:E,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),{ML}!E:E,"<="&TODAY(),{ML}!O:O,"Yes")',
     "Paid this month", GREEN, None),
    (17, "TODAY CONVERTED",
     f'=COUNTIFS({ML}!E:E,TODAY(),{ML}!F:F,"Converted")',
     "Converted today", AMBER, None),
    (22, "TOTAL LEADS",
     f"=COUNTA({ML}!B:B)-1",
     "All leads ever", DGRAY, None),
]

for col, title, formula, sublabel, accent, numfmt in kpis7_r1:
    kpi_card(ws7, 6, col, title, formula, sublabel, accent, numfmt=numfmt)
for col, title, formula, sublabel, accent, numfmt in kpis7_r2:
    kpi_card(ws7, 10, col, title, formula, sublabel, accent, numfmt=numfmt)

# Helper: daily 14 days trend
hws7.cell(1,1).value="Date"; hws7.cell(1,2).value="Leads"
for i in range(14):
    offset = 13 - i
    hws7.cell(2+i, 1).value = f"=TODAY()-{offset}"
    hws7.cell(2+i, 1).number_format = "DD-MMM"
    hws7.cell(2+i, 2).value = f'=COUNTIF({ML}!E:E,TODAY()-{offset})'

# Helper: weekly trend 8 weeks
hws7.cell(17,1).value="Week Ending"; hws7.cell(17,2).value="Leads"
for i in range(8):
    offset_end = (7-i-1)*7
    offset_start = offset_end + 6
    hws7.cell(18+i, 1).value = f"=TODAY()-{offset_end}"
    hws7.cell(18+i, 1).number_format = "DD-MMM"
    hws7.cell(18+i, 2).value = f'=COUNTIFS({ML}!E:E,">="&TODAY()-{offset_start},{ML}!E:E,"<="&TODAY()-{offset_end})'

# Charts
section_hdr(ws7, 14, 2, 30, "  TREND CHARTS   —   Daily, Weekly, and Monthly patterns")
ws7.row_dimensions[14].height = 26

# Daily line
line7a = LineChart()
line7a.title="Daily Leads — Last 14 Days"; line7a.style=10; line7a.legend=None
line7a.y_axis.title="Leads"; line7a.x_axis.title="Date"
c7a = Reference(hws7, min_col=1, max_col=1, min_row=2, max_row=15)
v7a = Reference(hws7, min_col=2, max_col=2, min_row=1, max_row=15)
line7a.add_data(v7a, titles_from_data=True)
line7a.set_categories(c7a)
line7a.series[0].graphicalProperties.line.solidFill = SKY
line7a.series[0].graphicalProperties.line.width = 28000
line7a.series[0].marker.symbol = "circle"; line7a.series[0].marker.size = 6
line7a.series[0].smooth = True
line7a.width = 22; line7a.height = 10
ws7.add_chart(line7a, "B15")

# Weekly trend
line7b = LineChart()
line7b.title="Weekly Trend — Last 8 Weeks"; line7b.style=10; line7b.legend=None
line7b.y_axis.title="Leads"; line7b.x_axis.title="Week Ending"
c7b = Reference(hws7, min_col=1, max_col=1, min_row=18, max_row=25)
v7b = Reference(hws7, min_col=2, max_col=2, min_row=17, max_row=25)
line7b.add_data(v7b, titles_from_data=True)
line7b.set_categories(c7b)
line7b.series[0].graphicalProperties.line.solidFill = NAVY
line7b.series[0].graphicalProperties.line.width = 28000
line7b.series[0].marker.symbol = "square"; line7b.series[0].marker.size = 6
line7b.series[0].smooth = True
line7b.width = 22; line7b.height = 10
ws7.add_chart(line7b, "B35")

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  DASHBOARD 8 – STATUS DASHBOARD                                         ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
H8 = "_DB8_Data"
if H8 in wb.sheetnames: del wb[H8]
hws8 = wb.create_sheet(H8); hws8.sheet_state = "hidden"

D8 = "Status Dashboard"
if D8 in wb.sheetnames: del wb[D8]
ws8 = wb.create_sheet(D8, 7)

setup_sheet(ws8, PURPLE,
    "  STATUS DASHBOARD   |   Lead Pipeline Breakdown by Status",
    "  Track where every lead stands in the sales funnel  \u00b7  Auto-updates",
    PURPLE, "6D28D9", "DDD6FE")

STATUSES = ["New","Contacted","Interested","Converted","No Answer",
            "Not Interested","Follow-up","Second Call Pending","Second Call Completed"]
STATUS_COLORS = [NAVY, BLUE, ROYAL, GREEN, AMBER, RED, ORANGE, PURPLE, TEAL]
STATUS_LIGHT   = [LINDIGO, "DBEAFE", "E0E7FF", LGREEN, LAMBER, LRED, LORANGE, LPURPLE, LTEAL]

# Helper data
hws8.cell(1,1).value="Status"; hws8.cell(1,2).value="Count"
for i, s in enumerate(STATUSES):
    hws8.cell(2+i, 1).value = s
    hws8.cell(2+i, 2).value = f'=COUNTIF({ML}!F:F,"{s}")'
# Second call done / not done
hws8.cell(12,1).value="Second Call Done"; hws8.cell(12,2).value=f'=COUNTIF({ML}!I:I,"Yes")'
hws8.cell(13,1).value="Second Call Pending"; hws8.cell(13,2).value=f'=COUNTIF({ML}!I:I,"No")'
hws8.cell(14,1).value="Second Call Not Set"; hws8.cell(14,2).value=f'=COUNTBLANK({ML}!I2:I10000)'

# Status Cards
section_hdr(ws8, 5, 2, 30, "  STATUS KPI CARDS   —   Live count per status")
ws8.row_dimensions[5].height = 26

for i, (status, accent) in enumerate(zip(STATUSES, STATUS_COLORS)):
    row = 6 if i < 5 else 10
    col = 2 + (i % 5) * 5
    if i == 5:
        for r in range(9,13): ws8.row_dimensions[r].height = 14
        ws8.row_dimensions[9].height = 10
    kpi_card(ws8, row, col, status, f'=COUNTIF({ML}!F:F,"{status}")',
             f"Leads with status = {status}", accent, W=5)

for r in [6,7,8,10,11,12]: ws8.row_dimensions[r].height = 14
ws8.row_dimensions[7].height = 38; ws8.row_dimensions[11].height = 38
ws8.row_dimensions[8].height = 15; ws8.row_dimensions[12].height = 15

# Status funnel table
section_hdr(ws8, 14, 2, 30, "  FUNNEL VIEW   —   Lead progression pipeline")
ws8.row_dimensions[14].height = 26; ws8.row_dimensions[15].height = 22

merge_set(ws8, 15, 2, 15, 10, val="Status Stage", fill_color=NAVY,
          font=ft(9, True, WHITE), align=al("center", "center"))
merge_set(ws8, 15, 11, 15, 18, val="Count", fill_color=BLUE,
          font=ft(9, True, WHITE), align=al("center", "center"))
merge_set(ws8, 15, 19, 15, 26, val="% of Total", fill_color=TEAL,
          font=ft(9, True, WHITE), align=al("center", "center"))

for i, (status, scolor, lcolor) in enumerate(zip(STATUSES, STATUS_COLORS, STATUS_LIGHT)):
    r = 16 + i
    ws8.row_dimensions[r].height = 20
    merge_set(ws8, r, 2, r, 10, val=status, fill_color=scolor,
              font=ft(10, True, WHITE), align=al("left", "center"))
    cnt_f = f'=COUNTIF({ML}!F:F,"{status}")'
    merge_set(ws8, r, 11, r, 18, val=cnt_f, fill_color=lcolor,
              font=ft(11, True, scolor), align=al("center", "center"))
    pct_c = merge_set(ws8, r, 19, r, 26,
              val=f'=IFERROR(COUNTIF({ML}!F:F,"{status}")/(COUNTA({ML}!B:B)-1),0)',
              fill_color=lcolor, font=ft(10, False, DGRAY), align=al("center", "center"))
    pct_c.number_format = "0.0%"

# Charts
chart_r8 = 16 + len(STATUSES) + 2
ws8.row_dimensions[chart_r8].height = 26
section_hdr(ws8, chart_r8, 2, 30, "  STATUS CHARTS")

cat8 = Reference(hws8, min_col=1, max_col=1, min_row=2, max_row=10)
val8 = Reference(hws8, min_col=2, max_col=2, min_row=1, max_row=10)

# Pie
pie8 = PieChart(); pie8.title = "Lead Status Distribution"; pie8.style = 10
pie8.add_data(val8, titles_from_data=True); pie8.set_categories(cat8)
for idx, color in enumerate(STATUS_COLORS):
    pt = DataPoint(idx=idx); pt.graphicalProperties.solidFill = color
    pie8.series[0].dPt.append(pt)
pie8.width = 12; pie8.height = 11
ws8.add_chart(pie8, "B" + str(chart_r8+1))

# Bar
bar8 = BarChart(); bar8.type = "bar"; bar8.title = "Status Count"; bar8.style = 10
bar8.legend = None; bar8.y_axis.title = "Status"; bar8.x_axis.title = "Count"
bar8.add_data(val8, titles_from_data=True); bar8.set_categories(cat8)
bar8.series[0].graphicalProperties.solidFill = PURPLE
bar8.width = 15; bar8.height = 11
ws8.add_chart(bar8, "L" + str(chart_r8+1))

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  REPEAT STUDENT DASHBOARD                                               ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
DR = "Repeat Students"
if DR in wb.sheetnames: del wb[DR]
wsr = wb.create_sheet(DR, 8)
setup_sheet(wsr, INDIGO,
    "  REPEAT STUDENT DASHBOARD   |   Students who Enquired Multiple Times",
    "  Track returning students across different campaigns  \u00b7  Uses col M (Repeat Student?)",
    INDIGO, "3730A3", "C7D2FE")

section_hdr(wsr, 5, 2, 28, "  REPEAT STUDENT KPIs")
wsr.row_dimensions[5].height = 26
for r in [6,7,8]: wsr.row_dimensions[r].height = 14
wsr.row_dimensions[7].height = 38; wsr.row_dimensions[8].height = 15

kpis_r = [
    (2,  "REPEAT STUDENTS", f'=COUNTIF({ML}!M:M,"Yes*")', "Enquired before (col M = Yes)", INDIGO, None),
    (7,  "NEW STUDENTS",    f'=COUNTIF({ML}!M:M,"No")',   "First-time enquiries", GREEN, None),
    (12, "REPEAT RATE",
     f'=IFERROR(COUNTIF({ML}!M:M,"Yes*")/(COUNTA({ML}!B:B)-1),0)',
     "Repeat / Total leads", PURPLE, "0.0%"),
    (17, "REPEAT & PAID",
     f'=COUNTIFS({ML}!M:M,"Yes*",{ML}!O:O,"Yes")',
     "Repeat students who also paid", TEAL, None),
    (22, "REPEAT & CONVERTED",
     f'=COUNTIFS({ML}!M:M,"Yes*",{ML}!F:F,"Converted")',
     "Repeat students converted", AMBER, None),
]
for col, title, formula, sublabel, accent, numfmt in kpis_r:
    kpi_card(wsr, 6, col, title, formula, sublabel, accent, numfmt=numfmt)

# Repeat by Campaign table
section_hdr(wsr, 10, 2, 28, "  REPEAT BY CAMPAIGN & MEMBER")
wsr.row_dimensions[10].height = 26; wsr.row_dimensions[11].height = 22

campaigns_short = ["B1","B2","B3","B4","B5","B6","B7","B8"]
hdrs_r = [("Campaign", 4, INDIGO), ("Total Leads", 4, NAVY), ("Repeat", 4, PURPLE), ("Repeat %", 4, TEAL)]
col = 2
for ht, span, hcol in hdrs_r:
    merge_set(wsr, 11, col, 11, col+span-1, val=ht, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

for i, camp in enumerate(campaigns_short):
    r = 12 + i
    wsr.row_dimensions[r].height = 18
    bg = LINDIGO if i % 2 == 0 else WHITE
    col = 2
    merge_set(wsr, r, col, r, col+3, val=camp, fill_color=INDIGO,
              font=ft(10, True, WHITE), align=al("center", "center")); col += 4
    merge_set(wsr, r, col, r, col+3,
              val=f'=COUNTIF({ML}!L:L,"{camp}")', fill_color=bg,
              font=ft(10, True, DGRAY), align=al("center", "center")); col += 4
    merge_set(wsr, r, col, r, col+3,
              val=f'=COUNTIFS({ML}!L:L,"{camp}",{ML}!M:M,"Yes*")', fill_color=bg,
              font=ft(10, True, PURPLE), align=al("center", "center")); col += 4
    pct = merge_set(wsr, r, col, r, col+3,
              val=f'=IFERROR(COUNTIFS({ML}!L:L,"{camp}",{ML}!M:M,"Yes*")/COUNTIF({ML}!L:L,"{camp}"),0)',
              fill_color=bg, font=ft(10, False, DGRAY), align=al("center", "center"))
    pct.number_format = "0.0%"

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  FOLLOW-UP DASHBOARD                                                    ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
DF = "Follow-up Dashboard"
if DF in wb.sheetnames: del wb[DF]
wsf = wb.create_sheet(DF, 9)
setup_sheet(wsf, ORANGE,
    "  FOLLOW-UP DASHBOARD   |   Second Calls, Pending Follow-ups & Rates",
    "  Track which leads need a second call and who completed it  \u00b7  Based on col I (Second Call Done)",
    ORANGE, AMBER, "FED7AA")

section_hdr(wsf, 5, 2, 28, "  FOLLOW-UP KPIs   —   Based on Second Call Done column (col I)")
wsf.row_dimensions[5].height = 26
for r in [6,7,8,9,10,11,12]: wsf.row_dimensions[r].height = 14
wsf.row_dimensions[7].height = 38; wsf.row_dimensions[8].height = 15
wsf.row_dimensions[11].height = 38; wsf.row_dimensions[12].height = 15

kpis_f1 = [
    (2,  "SECOND CALL DONE",   f'=COUNTIF({ML}!I:I,"Yes")',  "Completed second calls", GREEN, None),
    (7,  "SECOND CALL PENDING",f'=COUNTIF({ML}!I:I,"No")',   "Need second call", AMBER, None),
    (12, "NOT SET",            f'=COUNTBLANK({ML}!I2:I10000)', "Second call not recorded", GRAY, None),
    (17, "FOLLOW-UP RATE",
     f'=IFERROR(COUNTIF({ML}!I:I,"Yes")/(COUNTIF({ML}!I:I,"Yes")+COUNTIF({ML}!I:I,"No")),0)',
     "Done / (Done + Pending)", TEAL, "0.0%"),
    (22, "TOTAL LEADS",        f"=COUNTA({ML}!B:B)-1", "All leads in system", NAVY, None),
]

kpis_f2 = [
    (2,  "2ND CALL + CONVERTED",
     f'=COUNTIFS({ML}!I:I,"Yes",{ML}!F:F,"Converted")',
     "Second call done & converted", GREEN, None),
    (7,  "2ND CALL + PAID",
     f'=COUNTIFS({ML}!I:I,"Yes",{ML}!O:O,"Yes")',
     "Second call done & paid", TEAL, None),
    (12, "PENDING + INTERESTED",
     f'=COUNTIFS({ML}!I:I,"No",{ML}!F:F,"Interested")',
     "Follow-up needed + interested", AMBER, None),
    (17, "PENDING + NO ANSWER",
     f'=COUNTIFS({ML}!I:I,"No",{ML}!F:F,"No Answer")',
     "Pending call, no answer last time", RED, None),
]

for col, title, formula, sublabel, accent, numfmt in kpis_f1:
    kpi_card(wsf, 6, col, title, formula, sublabel, accent, numfmt=numfmt)
for col, title, formula, sublabel, accent, numfmt in kpis_f2:
    kpi_card(wsf, 10, col, title, formula, sublabel, accent, numfmt=numfmt)

# Per-member follow-up table
section_hdr(wsf, 14, 2, 28, "  FOLLOW-UP BY MEMBER   —   Who has pending second calls?")
wsf.row_dimensions[14].height = 26; wsf.row_dimensions[15].height = 22

hdrs_f = [("Member", 5, NAVY), ("2nd Done", 4, GREEN), ("2nd Pending", 4, AMBER),
          ("Not Set", 4, GRAY), ("Done %", 4, TEAL)]
col = 2
for ht, span, hcol in hdrs_f:
    merge_set(wsf, 15, col, 15, col+span-1, val=ht, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

for mi, m in enumerate(members):
    r = 16 + mi
    wsf.row_dimensions[r].height = 18
    bg = LAMBER if mi % 2 == 0 else WHITE
    acc = MEMBER_ACCENT_COLORS[mi % len(MEMBER_ACCENT_COLORS)]
    col = 2
    merge_set(wsf, r, col, r, col+4, val=m, fill_color=acc,
              font=ft(10, True, WHITE), align=al("left", "center")); col += 5
    merge_set(wsf, r, col, r, col+3,
              val=f'=COUNTIFS({ML}!D:D,"{m}",{ML}!I:I,"Yes")', fill_color=bg,
              font=ft(10, True, GREEN), align=al("center", "center")); col += 4
    merge_set(wsf, r, col, r, col+3,
              val=f'=COUNTIFS({ML}!D:D,"{m}",{ML}!I:I,"No")', fill_color=bg,
              font=ft(10, True, AMBER), align=al("center", "center")); col += 4
    merge_set(wsf, r, col, r, col+3,
              val=f'=COUNTIFS({ML}!D:D,"{m}",{ML}!I:I,"")', fill_color=bg,
              font=ft(10, False, GRAY), align=al("center", "center")); col += 4
    pct = merge_set(wsf, r, col, r, col+3,
              val=f'=IFERROR(COUNTIFS({ML}!D:D,"{m}",{ML}!I:I,"Yes")/(COUNTIFS({ML}!D:D,"{m}",{ML}!I:I,"Yes")+COUNTIFS({ML}!D:D,"{m}",{ML}!I:I,"No")),0)',
              fill_color=bg, font=ft(10, False, DGRAY), align=al("center", "center"))
    pct.number_format = "0.0%"

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  REVENUE DASHBOARD (FUTURE READY SHELL)                                 ║
# ╚═══════════════════════════════════════════════════════════════════════════╝
DRV = "Revenue Dashboard"
if DRV in wb.sheetnames: del wb[DRV]
wsrv = wb.create_sheet(DRV, 10)
setup_sheet(wsrv, TEAL,
    "  REVENUE DASHBOARD   |   Future-Ready Revenue Tracking Shell",
    "  Add a 'Fee' column (col Q) to Master Leads to activate all formulas below",
    TEAL, GREEN, "A7F3D0")

# Notice banner
for r in range(5,9): wsrv.row_dimensions[r].height = 14
wsrv.row_dimensions[6].height = 40

for c in range(1, 30): wsrv.cell(5, c).fill = pf(AMBER)
merge_set(wsrv, 5, 2, 5, 28,
    val="  ACTION REQUIRED: Add a column Q (Fee / Amount) to Master Leads to activate revenue formulas. All formulas below reference {ML}!Q:Q",
    fill_color=AMBER, font=ft(10, True, WHITE), align=al("left", "center"))

# Revenue KPI cards (dormant until fee column added)
fee_col = f"{ML}!Q:Q"
section_hdr(wsrv, 7, 2, 28, "  REVENUE KPIs   —   Will activate when Fee column (Q) is added to Master Leads")
wsrv.row_dimensions[7].height = 26
for r in [8,9,10]: wsrv.row_dimensions[r].height = 14
wsrv.row_dimensions[9].height = 38; wsrv.row_dimensions[10].height = 15

kpis_rv = [
    (2,  "TOTAL REVENUE",   f"=IFERROR(SUM({fee_col}),0)", "Sum of all fees paid", GREEN, '#,##0'),
    (8,  "AVG FEE",         f"=IFERROR(AVERAGEIF({fee_col},\">0\"),0)", "Average fee per student", TEAL, '#,##0'),
    (14, "REVENUE THIS MONTH",
     f'=IFERROR(SUMIFS({fee_col},{ML}!E:E,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),{ML}!E:E,"<="&TODAY()),0)',
     "This month's total revenue", ROYAL, '#,##0'),
    (20, "HIGHEST FEE",     f"=IFERROR(MAX({fee_col}),0)", "Highest single fee paid", PURPLE, '#,##0'),
]
for col, title, formula, sublabel, accent, numfmt in kpis_rv:
    vc = kpi_card(wsrv, 8, col, title, formula, sublabel, accent, W=5, numfmt=numfmt)

# Revenue breakdown tables
section_hdr(wsrv, 12, 2, 28, "  REVENUE BY GRADE   —   Add Fee col (Q) to activate")
wsrv.row_dimensions[12].height = 26; wsrv.row_dimensions[13].height = 22

hdrs_rv = [("Grade", 4, NAVY), ("Students Paid", 5, GREEN), ("Total Revenue", 5, TEAL), ("Avg Fee", 5, PURPLE)]
col = 2
for ht, span, hcol in hdrs_rv:
    merge_set(wsrv, 13, col, 13, col+span-1, val=ht, fill_color=hcol,
              font=ft(9, True, WHITE), align=al("center", "center"))
    col += span

for gi, g in enumerate(GRADES):
    r = 14 + gi
    wsrv.row_dimensions[r].height = 18
    bg = LTEAL if gi % 2 == 0 else WHITE
    from openpyxl.utils import get_column_letter as gcl
    col = 2
    merge_set(wsrv, r, col, r, col+3, val=f"Grade {g}", fill_color=TEAL,
              font=ft(10, True, WHITE), align=al("center", "center")); col += 4
    merge_set(wsrv, r, col, r, col+4,
              val=f'=COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes")', fill_color=bg,
              font=ft(10, True, GREEN), align=al("center", "center")); col += 5
    c_rev = merge_set(wsrv, r, col, r, col+4,
              val=f'=IFERROR(SUMIF({ML}!G:G,{g},{fee_col}),0)', fill_color=bg,
              font=ft(10, True, TEAL), align=al("center", "center"))
    c_rev.number_format = '#,##0'; col += 5
    c_avg = merge_set(wsrv, r, col, r, col+4,
              val=f'=IFERROR(SUMIF({ML}!G:G,{g},{fee_col})/COUNTIFS({ML}!G:G,{g},{ML}!O:O,"Yes"),0)',
              fill_color=bg, font=ft(10, False, DGRAY), align=al("center", "center"))
    c_avg.number_format = '#,##0'

# ═══════════════════════════════════════════════════════════════════════════════
# REORDER SHEETS
# ═══════════════════════════════════════════════════════════════════════════════
desired_order = [
    "Executive Dashboard", "Grade Interest", "Paid Dashboard",
    "Member Performance", "Member Paid", "Campaign Dashboard",
    "Daily Dashboard", "Status Dashboard", "Repeat Students",
    "Follow-up Dashboard", "Revenue Dashboard",
    "Instructions - READ FIRST", "Members", "Master Leads", "Leads Summary",
    "_DB1_Data", "_DB2_Data", "_DB3_Data", "_DB4_Data", "_DB5_Data",
    "_DB6_Data", "_DB7_Data", "_DB8_Data"
]
current = {s: i for i, s in enumerate(wb.sheetnames)}
for pos, name in enumerate(desired_order):
    if name in current:
        wb.move_sheet(name, offset=pos - wb.index(wb[name]))

wb.save(DST)
print("All remaining dashboards saved!")
print("Final sheets:", wb.sheetnames)
