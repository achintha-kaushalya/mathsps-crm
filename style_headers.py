#!/usr/bin/env python3
"""
Sets custom header colors on the Master Leads sheet of Master_Leads_GoogleSheets.xlsx
to match the user's manual colors:
- Col A, B, C (Phone Normalized, Raw Phone, F-Code): Navy Blue (0F172A)
- Col D, E, F (Assigned Member, Date Added, Status): Blue (2563EB)
- Col G, H (Grade, Comments): Slate Gray (475569)
- Col I, J (Campaign, Repeat Student): Mustard Gold (D97706)
- Col K, L (Duplicate Check, Previous F-Code): Rust Red (B91C1C)
- Col M, N (Second Call Done, Second Call Notes): Forest Green (15803D)
- Col O, P (Paid, Grade Final/Interested): Bright Blue (0284C7)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DST = r'c:\Users\achir\Desktop\Lead tracking PS\Master_Leads_GoogleSheets.xlsx'

wb = openpyxl.load_workbook(DST)
ws = wb['Master Leads']

# Define colors
colors = {
    "navy": "0D2744",       # A, B, C
    "blue": "1E40AF",       # D, E, F
    "slate": "475569",      # G, H
    "gold": "D97706",       # I, J
    "red": "B91C1C",        # K, L
    "green": "166534",      # M, N
    "sky": "0369A1"         # O, P
}

color_mapping = [
    ("navy", [1, 2, 3]),
    ("blue", [4, 5, 6]),
    ("slate", [7, 8]),
    ("gold", [9, 10]),
    ("red", [11, 12]),
    ("green", [13, 14]),
    ("sky", [15, 16])
]

# Set header styles
for group_color, cols in color_mapping:
    hex_color = colors[group_color]
    fill = PatternFill("solid", fgColor=hex_color)
    font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="medium", color="000000"))
    
    for c in cols:
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

wb.save(DST)
print("Applied column header colors to Master Leads workbook successfully!")
